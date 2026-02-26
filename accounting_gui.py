#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
家纺四件套记账系统 - GUI版本
功能：图形化界面实时记账
作者：AI Assistant
日期：2026-02-06
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import csv
import urllib.request
import urllib.error
import threading
import webbrowser
import sys
import subprocess
import time

# 导入打印模块
from receipt_printer import ReceiptPrinter, get_printer_list

# 导入自动更新模块
try:
    from auto_updater import check_for_updates, perform_update, get_current_version
except ImportError:
    def check_for_updates(silent=True):
        return False, None, "1.12.0", ""
    def perform_update(callback=None):
        return False, "更新模块未安装"
    def get_current_version():
        return "1.13.0"

# 版本信息
VERSION = "1.14.0"
GITHUB_REPO = "andyyuzy-76/textile-accounting"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"

class AccountingApp:
    # 深蓝色系配色方案
    COLORS = {
        'primary': '#1E3A5F',       # 深蓝主色
        'primary_light': '#2C5282', # 浅蓝
        'primary_dark': '#1A365D',  # 更深蓝
        'secondary': '#4A5568',     # 次要色
        'success': '#276749',       # 深绿
        'warning': '#C05621',       # 深橙
        'danger': '#C53030',        # 深红
        'dark': '#1A202C',          # 深色文字
        'gray': '#4A5568',          # 灰色文字
        'light': '#EDF2F7',         # 浅灰背景
        'white': '#FFFFFF',         # 白色
        'card_bg': '#FFFFFF',       # 卡片背景
        'border': '#CBD5E0',        # 边框色
        'divider': '#A0AEC0',       # 分割线
        'text': '#2D3748',          # 主文字
        'text_light': '#718096',    # 次要文字
        'text_hint': '#A0AEC0',     # 提示文字
        'header_bg': '#1E3A5F',     # 标题背景-深蓝
        'hover_bg': '#EBF8FF',      # 悬停背景
        'selected_bg': '#BEE3F8',   # 选中背景
        'input_bg': '#F7FAFC',      # 输入框背景
    }
    
    def __init__(self, root):
        self.root = root
        self.root.title(f"🏠 家纺四件套记账系统 v{VERSION}")
        self.root.geometry("1100x750")
        self.root.configure(bg=self.COLORS['light'])
        self.root.state('zoomed')  # 窗口最大化
        
        # 数据文件路径
        home_dir = os.path.expanduser("~")
        self.data_dir = os.path.join(home_dir, ".accounting-tool")
        self.data_file = os.path.join(self.data_dir, "records.json")
        os.makedirs(self.data_dir, exist_ok=True)
        
        # 加载数据
        self.records = self.load_records()

        # 显示模式：True=只显示今天，False=显示全部
        self.showing_today_only = True

        # 初始化打印机
        self.receipt_printer = ReceiptPrinter()
        self.load_printer_settings()
        
        # 配置ttk样式
        self.setup_styles()

        # 创建界面
        self.create_widgets()
        
        # 刷新显示（默认只显示今天）
        self.refresh_display()
        
        # 绑定快捷键（F5刷新，Ctrl+Enter添加记录）
        self.root.bind('<F5>', lambda e: self.refresh_display())
        self.root.bind('<Control-Return>', lambda e: self.add_record())
    
    def setup_styles(self):
        """配置ttk样式 - 深蓝色风格"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Treeview样式 - 深蓝色表头
        style.configure('Custom.Treeview',
                       background=self.COLORS['white'],
                       foreground=self.COLORS['text'],
                       fieldbackground=self.COLORS['white'],
                       borderwidth=0,
                       rowheight=28,
                       font=('微软雅黑', 10))
        style.configure('Custom.Treeview.Heading',
                       font=('微软雅黑', 10, 'bold'),
                       background=self.COLORS['primary'],
                       foreground=self.COLORS['white'])
        style.map('Custom.Treeview',
                 background=[('selected', self.COLORS['selected_bg'])],
                 foreground=[('selected', self.COLORS['primary'])])
        
        # Combobox样式
        style.configure('Custom.TCombobox',
                       fieldbackground=self.COLORS['white'],
                       background=self.COLORS['white'],
                       borderwidth=1)
    
    def load_records(self) -> List[Dict]:
        """加载历史记录"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def save_records(self):
        """保存记录"""
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(self.records, f, ensure_ascii=False, indent=2)

    def load_printer_settings(self):
        """加载打印机设置"""
        settings_file = os.path.join(self.data_dir, "printer_settings.json")
        self.printer_settings = {
            'shop_name': '家纺四件套',
            'shop_address': '',
            'shop_phone': '',
            'footer_text': '谢谢惠顾，欢迎下次光临！',
            'printer_name': '',  # 空字符串表示使用默认打印机
            'auto_print': False,  # 是否自动打印
            'paper_width': 58,  # 纸张宽度：58mm或80mm
            'compact_mode': True  # 紧凑模式：一张纸打印（推荐58mm）
        }

        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    loaded_settings = json.load(f)
                    self.printer_settings.update(loaded_settings)

                    # 应用到打印机
                    self.receipt_printer.set_shop_info(
                        name=self.printer_settings['shop_name'],
                        address=self.printer_settings['shop_address'],
                        phone=self.printer_settings['shop_phone']
                    )
                    self.receipt_printer.footer_text = self.printer_settings['footer_text']
                    self.receipt_printer.receipt_width = 32 if self.printer_settings['paper_width'] == 58 else (44 if self.printer_settings['paper_width'] == 76 else 48)
            except:
                pass

    def save_printer_settings(self, settings):
        """保存打印机设置"""
        settings_file = os.path.join(self.data_dir, "printer_settings.json")
        try:
            # 更新内存中的设置
            self.printer_settings.update(settings)

            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(self.printer_settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("错误", f"保存设置失败: {str(e)}")
    
    def create_widgets(self):
        """创建界面组件"""
        # ===== 标题栏 - 深蓝色 =====
        title_frame = tk.Frame(self.root, bg=self.COLORS['primary'], height=56)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        # 标题文字
        title_label = tk.Label(
            title_frame, 
            text="🏠 家纺记账", 
            font=('微软雅黑', 18, 'bold'),
            bg=self.COLORS['primary'],
            fg=self.COLORS['white']
        )
        title_label.pack(pady=12)
        
        # 版本标签
        version_label = tk.Label(
            title_frame,
            text=f"v{VERSION}",
            font=('微软雅黑', 9),
            bg=self.COLORS['primary'],
            fg='#A0C4E8'
        )
        version_label.place(relx=0.96, rely=0.5, anchor='e')
        
        # 分割线
        tk.Frame(self.root, bg=self.COLORS['divider'], height=1).pack(fill=tk.X)
        
        # ===== 主内容区 =====
        main_frame = tk.Frame(self.root, bg=self.COLORS['light'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # ===== 左侧录入区 =====
        left_frame = tk.Frame(main_frame, bg=self.COLORS['card_bg'], 
                             highlightbackground=self.COLORS['border'], 
                             highlightthickness=1)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
        
        # 录入区标题
        header_frame = tk.Frame(left_frame, bg=self.COLORS['primary_light'], height=40)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame, 
            text="📝 新记录", 
            font=('微软雅黑', 12, 'bold'),
            bg=self.COLORS['primary_light'],
            fg=self.COLORS['white']
        ).pack(pady=8, padx=12, anchor='w')
        
        # 录入表单容器
        form_frame = tk.Frame(left_frame, bg=self.COLORS['card_bg'])
        form_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # ===== 日期行 =====
        tk.Label(form_frame, text="日期", font=('微软雅黑', 11), 
                bg=self.COLORS['card_bg'], fg=self.COLORS['dark']).pack(anchor='w', pady=(0, 4))
        
        date_row = tk.Frame(form_frame, bg=self.COLORS['card_bg'])
        date_row.pack(fill=tk.X, pady=(0, 8))
        
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        date_entry = tk.Entry(date_row, textvariable=self.date_var, 
                             font=('微软雅黑', 11), width=16,
                             bg=self.COLORS['white'], fg=self.COLORS['dark'],
                             relief='solid', borderwidth=1,
                             highlightthickness=0)
        date_entry.pack(side=tk.LEFT, ipady=5)
        
        today_btn = tk.Button(date_row, text="今天", command=self.set_today, 
                             font=('微软雅黑', 10), bg=self.COLORS['light'],
                             fg=self.COLORS['primary'], relief='flat',
                             cursor='hand2', borderwidth=0)
        today_btn.pack(side=tk.LEFT, padx=8)
        
        # ===== 记录类型 =====
        tk.Label(form_frame, text="类型", font=('微软雅黑', 11), 
                bg=self.COLORS['card_bg'], fg=self.COLORS['dark']).pack(anchor='w', pady=(8, 4))
        
        type_frame = tk.Frame(form_frame, bg=self.COLORS['card_bg'])
        type_frame.pack(fill=tk.X, pady=(0, 8))
        
        self.record_type_var = tk.StringVar(value="sale")
        
        # 销售单选按钮
        sale_rb = tk.Radiobutton(type_frame, text="销售", variable=self.record_type_var, 
                                value="sale", font=('微软雅黑', 10), 
                                bg=self.COLORS['card_bg'], fg=self.COLORS['dark'],
                                selectcolor=self.COLORS['light'], 
                                activebackground=self.COLORS['card_bg'],
                                activeforeground=self.COLORS['success'])
        sale_rb.pack(side=tk.LEFT, padx=(0, 16))
        
        # 退货单选按钮
        return_rb = tk.Radiobutton(type_frame, text="退货", variable=self.record_type_var,
                                  value="return", font=('微软雅黑', 10), 
                                  bg=self.COLORS['card_bg'], fg=self.COLORS['dark'],
                                  selectcolor=self.COLORS['light'],
                                  activebackground=self.COLORS['card_bg'],
                                  activeforeground=self.COLORS['danger'])
        return_rb.pack(side=tk.LEFT)
        
        # ===== 商品明细区域 =====
        tk.Label(form_frame, text="商品明细", font=('微软雅黑', 11), 
                bg=self.COLORS['card_bg'], fg=self.COLORS['dark']).pack(anchor='w', pady=(8, 4))
        
        # 商品表格容器
        items_container_frame = tk.Frame(form_frame, bg=self.COLORS['white'],
                                        relief='solid', borderwidth=1)
        items_container_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 4))
        
        # 商品表格头部
        header_row = tk.Frame(items_container_frame, bg=self.COLORS['light'])
        header_row.pack(fill=tk.X, padx=8, pady=6)
        
        tk.Label(header_row, text="数量", font=('微软雅黑', 10), 
                bg=self.COLORS['light'], fg=self.COLORS['gray'],
                width=10, anchor='center').pack(side=tk.LEFT, padx=2)
        tk.Label(header_row, text="单价", font=('微软雅黑', 10), 
                bg=self.COLORS['light'], fg=self.COLORS['gray'],
                width=10, anchor='center').pack(side=tk.LEFT, padx=2)
        tk.Label(header_row, text="小计", font=('微软雅黑', 10), 
                bg=self.COLORS['light'], fg=self.COLORS['gray'],
                width=12, anchor='center').pack(side=tk.LEFT, padx=2)
        tk.Label(header_row, text="", bg=self.COLORS['light'], width=3).pack(side=tk.LEFT)
        
        # 商品行容器（可滚动）
        self.items_container = tk.Frame(items_container_frame, bg=self.COLORS['white'])
        self.items_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=2)
        
        # 存储商品行数据
        self.item_rows = []
        
        # 添加第一行
        self.add_item_row()
        
        # 添加商品行按钮
        add_row_btn = tk.Button(items_container_frame, text="+ 添加商品", 
                               command=self.add_item_row,
                               font=('微软雅黑', 10), bg=self.COLORS['white'],
                               fg=self.COLORS['primary'], relief='flat',
                               cursor='hand2', borderwidth=0,
                               activebackground=self.COLORS['light'])
        add_row_btn.pack(pady=6)
        
        # ===== 汇总区域 =====
        summary_frame = tk.Frame(form_frame, bg=self.COLORS['light'])
        summary_frame.pack(fill=tk.X, pady=8, ipady=6)
        
        tk.Label(summary_frame, text="汇总", font=('微软雅黑', 11), 
                bg=self.COLORS['light'], fg=self.COLORS['dark']).pack(side=tk.LEFT, padx=10)
        
        self.summary_qty_var = tk.StringVar(value="0套")
        tk.Label(summary_frame, textvariable=self.summary_qty_var, 
                font=('微软雅黑', 11), bg=self.COLORS['light'],
                fg=self.COLORS['dark']).pack(side=tk.LEFT, padx=10)
        
        self.summary_total_var = tk.StringVar(value="¥0.00")
        tk.Label(summary_frame, textvariable=self.summary_total_var, 
                font=('微软雅黑', 13, 'bold'), bg=self.COLORS['light'],
                fg=self.COLORS['primary']).pack(side=tk.LEFT, padx=10)
        
        # ===== 备注区域 =====
        tk.Label(form_frame, text="备注", font=('微软雅黑', 11), 
                bg=self.COLORS['card_bg'], fg=self.COLORS['dark']).pack(anchor='w', pady=(8, 4))
        
        self.note_text = tk.Text(form_frame, font=('微软雅黑', 10), 
                                width=30, height=2,
                                bg=self.COLORS['white'], fg=self.COLORS['dark'],
                                relief='solid', borderwidth=1,
                                highlightthickness=0)
        self.note_text.pack(fill=tk.X, ipady=4)
        self.note_text.bind('<Return>', self.on_note_return)
        
        # ===== 操作按钮区 =====
        btn_frame = tk.Frame(left_frame, bg=self.COLORS['card_bg'])
        btn_frame.pack(fill=tk.X, padx=12, pady=12)
        
        # 添加记录按钮
        add_btn = tk.Button(
            btn_frame,
            text="✅ 添加记录",
            command=self.add_record,
            font=('微软雅黑', 11, 'bold'),
            bg=self.COLORS['primary'],
            fg=self.COLORS['white'],
            height=2,
            relief='flat',
            cursor='hand2',
            borderwidth=0
        )
        add_btn.pack(fill=tk.X, pady=4)

        # 快捷提示
        tip_label = tk.Label(
            btn_frame,
            text="Ctrl+Enter 快速提交",
            font=('微软雅黑', 9),
            bg=self.COLORS['card_bg'],
            fg=self.COLORS['text_light']
        )
        tip_label.pack(pady=4)
        
        # 清空按钮
        clear_btn = tk.Button(
            btn_frame, 
            text="清空表单", 
            command=self.clear_form,
            font=('微软雅黑', 10),
            bg=self.COLORS['light'],
            fg=self.COLORS['text'],
            relief='flat',
            cursor='hand2',
            borderwidth=0
        )
        clear_btn.pack(fill=tk.X, pady=4)
        
        # ===== 今日统计卡片 =====
        stats_card = tk.Frame(left_frame, bg=self.COLORS['light'])
        stats_card.pack(fill=tk.X, padx=12, pady=(4, 12))
        
        stats_header = tk.Frame(stats_card, bg=self.COLORS['primary_light'], height=32)
        stats_header.pack(fill=tk.X)
        stats_header.pack_propagate(False)
        
        tk.Label(stats_header, text="📊 今日统计", font=('微软雅黑', 10, 'bold'),
                bg=self.COLORS['primary_light'], fg=self.COLORS['white']).pack(pady=5)
        
        self.stats_label = tk.Label(
            stats_card, 
            text="加载中...", 
            font=('微软雅黑', 10),
            bg=self.COLORS['light'],
            fg=self.COLORS['text'],
            justify=tk.LEFT
        )
        self.stats_label.pack(padx=10, pady=10)
        
        # ===== 右侧记录列表 =====
        right_frame = tk.Frame(main_frame, bg=self.COLORS['card_bg'],
                              highlightbackground=self.COLORS['border'],
                              highlightthickness=1)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # 记录列表标题
        list_header = tk.Frame(right_frame, bg=self.COLORS['primary_light'], height=40)
        list_header.pack(fill=tk.X)
        list_header.pack_propagate(False)
        
        tk.Label(list_header, text="📋 记录列表", font=('微软雅黑', 12, 'bold'),
                bg=self.COLORS['primary_light'], fg=self.COLORS['white']).pack(side=tk.LEFT, pady=8, padx=12)
        
        # ===== 筛选区 =====
        filter_frame = tk.Frame(right_frame, bg=self.COLORS['card_bg'])
        filter_frame.pack(fill=tk.X, padx=10, pady=8)

        # 显示状态标签
        self.status_label = tk.Label(filter_frame, text="今日", 
                                    font=('微软雅黑', 10, 'bold'), 
                                    bg=self.COLORS['card_bg'], fg=self.COLORS['primary'])
        self.status_label.pack(side=tk.LEFT, padx=4)

        tk.Label(filter_frame, text="|", font=('微软雅黑', 10), 
                bg=self.COLORS['card_bg'], fg=self.COLORS['divider']).pack(side=tk.LEFT, padx=6)

        # 日期选择器
        tk.Label(filter_frame, text="日期", font=('微软雅黑', 10), 
                bg=self.COLORS['card_bg'], fg=self.COLORS['gray']).pack(side=tk.LEFT, padx=(2, 4))

        # 年份选择
        self.filter_year_var = tk.StringVar(value=str(datetime.now().year))
        year_values = [str(y) for y in range(2020, 2031)]
        year_combo = ttk.Combobox(filter_frame, textvariable=self.filter_year_var, 
                                  values=year_values, width=6, 
                                  font=('微软雅黑', 10), state='readonly')
        year_combo.pack(side=tk.LEFT, padx=1)
        tk.Label(filter_frame, text="-", font=('微软雅黑', 10), 
                bg=self.COLORS['card_bg'], fg=self.COLORS['gray']).pack(side=tk.LEFT)

        # 月份选择
        self.filter_month_var = tk.StringVar(value=str(datetime.now().month).zfill(2))
        month_values = [str(m).zfill(2) for m in range(1, 13)]
        month_combo = ttk.Combobox(filter_frame, textvariable=self.filter_month_var,
                                   values=month_values, width=3,
                                   font=('微软雅黑', 10), state='readonly')
        month_combo.pack(side=tk.LEFT, padx=1)
        tk.Label(filter_frame, text="-", font=('微软雅黑', 10),
                bg=self.COLORS['card_bg'], fg=self.COLORS['gray']).pack(side=tk.LEFT)

        # 日期选择
        self.filter_day_var = tk.StringVar(value=str(datetime.now().day).zfill(2))
        day_values = [str(d).zfill(2) for d in range(1, 32)]
        day_combo = ttk.Combobox(filter_frame, textvariable=self.filter_day_var,
                                 values=day_values, width=3,
                                 font=('微软雅黑', 10), state='readonly')
        day_combo.pack(side=tk.LEFT, padx=1)

        # 确认按钮
        tk.Button(filter_frame, text="查询", command=self.confirm_date_filter,
                  font=('微软雅黑', 9), bg=self.COLORS['primary'],
                  fg=self.COLORS['white'], relief='flat',
                  cursor='hand2', borderwidth=0).pack(side=tk.LEFT, padx=8)

        # 快捷筛选按钮 - 深蓝风格
        tk.Button(filter_frame, text="今天", command=self.show_today_records,
                  font=('微软雅黑', 9), bg=self.COLORS['primary_light'],
                  fg=self.COLORS['white'], relief='flat',
                  cursor='hand2', borderwidth=0).pack(side=tk.LEFT, padx=3)
        tk.Button(filter_frame, text="本月", command=self.show_month_records,
                  font=('微软雅黑', 9), bg=self.COLORS['card_bg'],
                  fg=self.COLORS['text'], relief='solid', borderwidth=1,
                  cursor='hand2').pack(side=tk.LEFT, padx=3)
        tk.Button(filter_frame, text="本年", command=self.show_year_records,
                  font=('微软雅黑', 9), bg=self.COLORS['card_bg'],
                  fg=self.COLORS['text'], relief='solid', borderwidth=1,
                  cursor='hand2').pack(side=tk.LEFT, padx=3)
        tk.Button(filter_frame, text="全部", command=self.show_all_records,
                  font=('微软雅黑', 9), bg=self.COLORS['card_bg'],
                  fg=self.COLORS['text'], relief='solid', borderwidth=1,
                  cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        # ===== 记录表格 =====
        tree_frame = tk.Frame(right_frame, bg=self.COLORS['white'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        
        # 滚动条
        scrollbar_y = tk.Scrollbar(tree_frame, bg=self.COLORS['light'])
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        scrollbar_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, bg=self.COLORS['light'])
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 表格（树形显示：购买记录为父节点，退货为子节点）
        self.tree = ttk.Treeview(
            tree_frame,
            columns=('ID', '日期', '数量', '明细', '总金额', '备注'),
            show='tree headings',
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            style='Custom.Treeview'
        )
        
        # 设置树形列宽度
        self.tree.column('#0', width=30, stretch=False)
        
        # 设置列 - 深蓝风格表头
        self.tree.heading('ID', text='ID')
        self.tree.heading('日期', text='📅 日期')
        self.tree.heading('数量', text='📦 数量')
        self.tree.heading('明细', text='📋 明细')
        self.tree.heading('总金额', text='💵 金额')
        self.tree.heading('备注', text='📝 备注')
        
        self.tree.column('ID', width=45, anchor='center')
        self.tree.column('日期', width=90, anchor='center')
        self.tree.column('数量', width=60, anchor='center')
        self.tree.column('明细', width=160, anchor='w')
        self.tree.column('总金额', width=90, anchor='center')
        self.tree.column('备注', width=120, anchor='w')
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)

        # ===== 合计金额显示 =====
        total_frame = tk.Frame(right_frame, bg=self.COLORS['light'])
        total_frame.pack(fill=tk.X, padx=10, pady=(4, 10), ipady=6)

        tk.Label(total_frame, text="💰 合计:", font=('微软雅黑', 11, 'bold'), 
                bg=self.COLORS['light'], fg=self.COLORS['text']).pack(side=tk.LEFT, padx=10)
        self.total_label = tk.Label(total_frame, text="¥0.00", 
                                   font=('微软雅黑', 14, 'bold'),
                                   bg=self.COLORS['light'], fg=self.COLORS['primary'])
        self.total_label.pack(side=tk.LEFT, padx=4)

        # ===== 右键菜单 - 苹果风格 =====
        self.context_menu = tk.Menu(self.root, tearoff=0, 
                                   bg=self.COLORS['white'], fg=self.COLORS['dark'],
                                   activebackground=self.COLORS['selected_bg'],
                                   activeforeground=self.COLORS['dark'],
                                   font=('微软雅黑', 10))
        self.context_menu.add_command(label="查看明细", command=self.show_record_details)
        self.context_menu.add_command(label="打印小票", command=self.print_selected_record)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="编辑备注", command=self.edit_note)
        self.context_menu.add_command(label="编辑明细", command=self.edit_quantity_price)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="删除记录", command=self.delete_selected)
        self.tree.bind('<Button-3>', self.show_context_menu)
        self.tree.bind('<Double-Button-1>', self.show_record_details)  # 双击查看明细
        
        # ===== 底部按钮栏 - 深蓝色 =====
        bottom_frame = tk.Frame(self.root, bg=self.COLORS['primary'], height=50)
        bottom_frame.pack(fill=tk.X, side=tk.BOTTOM)
        bottom_frame.pack_propagate(False)
        
        btn_container = tk.Frame(bottom_frame, bg=self.COLORS['primary'])
        btn_container.pack(pady=8)
        
        # 创建按钮的辅助函数 - 深蓝风格
        def create_btn(parent, text, command, bg_color, fg_color=None):
            return tk.Button(
                parent,
                text=text,
                command=command,
                font=('微软雅黑', 10),
                bg=bg_color,
                fg=fg_color or self.COLORS['white'],
                relief='flat',
                cursor='hand2',
                padx=12,
                borderwidth=0,
                activebackground=self.COLORS['primary_dark'],
                activeforeground=self.COLORS['white']
            )
        
        create_btn(btn_container, "💾 导出CSV", self.export_csv, 
                  self.COLORS['primary_light']).pack(side=tk.LEFT, padx=4)
        create_btn(btn_container, "📥 导入CSV", self.import_csv,
                  self.COLORS['primary_light']).pack(side=tk.LEFT, padx=4)
        create_btn(btn_container, "📥 导入Excel", self.import_excel,
                  self.COLORS['primary_light']).pack(side=tk.LEFT, padx=4)
        create_btn(btn_container, "📊 月度统计", self.show_monthly_stats,
                  self.COLORS['warning']).pack(side=tk.LEFT, padx=4)
        create_btn(btn_container, "🖨️ 打印设置", self.show_printer_settings,
                  self.COLORS['secondary']).pack(side=tk.LEFT, padx=4)
        create_btn(btn_container, "⚙️ 系统设置", self.show_settings,
                  self.COLORS['secondary']).pack(side=tk.LEFT, padx=4)
        create_btn(btn_container, "❌ 退出", self.root.quit,
                  self.COLORS['danger']).pack(side=tk.LEFT, padx=4)

        # 设置初始焦点（第一个商品行的数量输入框）
        self.root.after(100, lambda: self.item_rows[0]['qty_entry'].focus_set() if self.item_rows else None)

    def add_item_row(self):
        """添加一个商品行"""
        row_frame = tk.Frame(self.items_container, bg=self.COLORS['white'])
        row_frame.pack(fill=tk.X, pady=4)
        
        qty_var = tk.StringVar()
        price_var = tk.StringVar()
        
        # 数量输入
        qty_entry = tk.Entry(row_frame, textvariable=qty_var, font=('微软雅黑', 10), 
                            width=10, bg=self.COLORS['white'], fg=self.COLORS['dark'],
                            relief='solid', borderwidth=1,
                            highlightthickness=0)
        qty_entry.pack(side=tk.LEFT, padx=2, ipady=3)
        qty_var.trace_add('write', lambda *args: self.update_item_subtotal(row_data))
        
        # 单价输入
        price_entry = tk.Entry(row_frame, textvariable=price_var, font=('微软雅黑', 10),
                              width=10, bg=self.COLORS['white'], fg=self.COLORS['dark'],
                              relief='solid', borderwidth=1,
                              highlightthickness=0)
        price_entry.pack(side=tk.LEFT, padx=2, ipady=3)
        price_var.trace_add('write', lambda *args: self.update_item_subtotal(row_data))
        
        # 小计显示
        subtotal_label = tk.Label(row_frame, text="¥0.00", font=('微软雅黑', 10),
                                 bg=self.COLORS['white'], fg=self.COLORS['gray'],
                                 width=12, anchor='w')
        subtotal_label.pack(side=tk.LEFT, padx=2)
        
        # 删除按钮 - 苹果风格文字按钮
        def delete_row():
            if len(self.item_rows) > 1:  # 至少保留一行
                row_frame.destroy()
                self.item_rows.remove(row_data)
                self.update_summary()
        
        del_btn = tk.Button(row_frame, text="×", command=delete_row, font=('微软雅黑', 12, 'bold'), 
                           bg=self.COLORS['white'], fg=self.COLORS['text_hint'], 
                           width=2, relief='flat', cursor='hand2',
                           borderwidth=0, activebackground=self.COLORS['light'])
        del_btn.pack(side=tk.LEFT, padx=2)
        
        row_data = {
            'qty_var': qty_var,
            'price_var': price_var,
            'subtotal_label': subtotal_label,
            'frame': row_frame,
            'qty_entry': qty_entry,
            'price_entry': price_entry
        }
        
        self.item_rows.append(row_data)
        
        # 绑定回车键跳转
        qty_entry.bind('<Return>', lambda e: price_entry.focus())
        price_entry.bind('<Return>', lambda e: self.on_price_enter(row_data))
        
        # 聚焦到数量输入框
        qty_entry.focus_set()
        
        return row_data
    
    def on_price_enter(self, current_row):
        """单价输入框回车：如果有值则添加新行，否则提交"""
        qty = current_row['qty_var'].get().strip()
        price = current_row['price_var'].get().strip()
        
        if qty and price:
            # 当前行有数据，添加新行
            new_row = self.add_item_row()
        else:
            # 当前行没数据，跳到备注或提交
            self.note_text.focus()
    
    def update_item_subtotal(self, row_data):
        """更新单行小计"""
        try:
            qty = int(row_data['qty_var'].get() or 0)
            price = float(row_data['price_var'].get() or 0)
            subtotal = qty * price
            row_data['subtotal_label'].config(text=f"¥{subtotal:.2f}")
        except:
            row_data['subtotal_label'].config(text="¥0.00")
        self.update_summary()
    
    def update_summary(self):
        """更新汇总信息"""
        total_qty = 0
        total_amount = 0.0
        
        for row in self.item_rows:
            try:
                qty = int(row['qty_var'].get() or 0)
                price = float(row['price_var'].get() or 0)
                total_qty += qty
                total_amount += qty * price
            except:
                pass
        
        self.summary_qty_var.set(f"{total_qty}套")
        self.summary_total_var.set(f"¥{total_amount:.2f}")

    def on_note_return(self, event):
        """备注框回车事件"""
        # Ctrl+Enter: 添加记录
        if event.state & 0x4:  # Ctrl键 (Windows)
            self.add_record()
            return 'break'
        # 普通回车: 插入换行
        return None  # 允许默认行为（换行）
    
    def insert_newline(self):
        """在备注框插入换行"""
        self.note_text.insert(tk.INSERT, '\n')
    
    def set_today(self):
        """设置日期为今天"""
        self.date_var.set(datetime.now().strftime("%Y-%m-%d"))
    
    def calculate_total(self, event=None):
        """计算总金额（兼容旧方法，现在通过update_summary实现）"""
        self.update_summary()
    
    def add_record(self):
        """添加记录（支持多商品行）"""
        try:
            date = self.date_var.get().strip()
            note = self.note_text.get('1.0', tk.END).strip()
            record_type = self.record_type_var.get()  # sale 或 return
            
            # 验证日期
            if not date:
                messagebox.showerror("错误", "请输入日期！")
                return
            
            # 收集所有商品行数据
            items = []
            total_quantity = 0
            total_amount = 0.0
            
            for row in self.item_rows:
                qty_str = row['qty_var'].get().strip()
                price_str = row['price_var'].get().strip()
                
                if qty_str and price_str:
                    try:
                        qty = int(qty_str)
                        price = float(price_str)
                        if qty > 0 and price > 0:
                            items.append({'quantity': qty, 'unit_price': price})
                            total_quantity += qty
                            total_amount += qty * price
                    except ValueError:
                        pass
            
            # 验证是否有有效商品
            if not items:
                messagebox.showerror("错误", "请至少添加一个有效的商品行（数量和单价都要大于0）！")
                return
            
            # 处理退货：数量为负数
            if record_type == "return":
                total_quantity = -total_quantity
                total_amount = -total_amount
                type_label = "退货"
                # 退货记录自动添加标识
                if note:
                    note = f"[退货] {note}"
                else:
                    note = "[退货]"
                # items中的数量也取负
                for item in items:
                    item['quantity'] = -item['quantity']
            else:
                type_label = "销售"
            
            # 计算平均单价（用于兼容旧数据格式显示）
            avg_price = abs(total_amount) / abs(total_quantity) if total_quantity != 0 else 0
            
            # 创建记录（保持向后兼容）
            record = {
                "id": len(self.records) + 1,
                "date": date,
                "quantity": total_quantity,
                "unit_price": avg_price,  # 平均单价，用于兼容
                "total_amount": total_amount,
                "note": note,
                "type": record_type,
                "items": items,  # 新增：商品明细
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            self.records.append(record)
            self.save_records()
            
            # 刷新显示
            self.refresh_display()
            self.clear_form()

            # 显示成功提示（带打印选项）
            abs_quantity = abs(total_quantity)
            items_count = len(items)
            success_msg = f"✅ {type_label}记录添加成功！\n日期: {date}\n商品: {items_count}种 共{abs_quantity}套\n金额: ¥{abs(total_amount):.2f}"
            self.show_success_message(success_msg, record)
            
        except ValueError as e:
            messagebox.showerror("错误", f"输入格式错误: {str(e)}")
    
    def show_success_message(self, message, record=None):
        """显示成功提示，带有打印选项"""
        popup = tk.Toplevel(self.root)
        popup.title("成功")
        popup.geometry("320x200")
        popup.transient(self.root)

        tk.Label(popup, text=message, font=('微软雅黑', 11), justify=tk.CENTER).pack(pady=15)

        # 按钮区域
        btn_frame = tk.Frame(popup)
        btn_frame.pack(pady=10)

        # 打印按钮（如果有记录）
        if record:
            print_btn = tk.Button(
                btn_frame,
                text="🖨️ 打印小票",
                command=lambda: [popup.destroy(), self.print_receipt(record)],
                font=('微软雅黑', 11),
                bg='#3498db',
                fg='white',
                width=12
            )
            print_btn.pack(side=tk.LEFT, padx=5)

            preview_btn = tk.Button(
                btn_frame,
                text="👁️ 预览",
                command=lambda: self.show_receipt_preview(record),
                font=('微软雅黑', 11),
                bg='#9b59b6',
                fg='white',
                width=10
            )
            preview_btn.pack(side=tk.LEFT, padx=5)

        # 确定按钮
        ok_btn = tk.Button(
            btn_frame,
            text="确定",
            command=popup.destroy,
            font=('微软雅黑', 11),
            bg='#27ae60',
            fg='white',
            width=10
        )
        if record:
            ok_btn.pack(side=tk.LEFT, padx=5)
        else:
            ok_btn.pack(pady=5)

        # 3秒后自动关闭（如果没有操作）
        popup.after(5000, popup.destroy)
    
    def clear_form(self):
        """清空表单"""
        # 清空备注
        self.note_text.delete('1.0', tk.END)
        
        # 清空所有商品行（保留第一行）
        while len(self.item_rows) > 1:
            row = self.item_rows.pop()
            row['frame'].destroy()
        
        # 清空第一行的数据
        if self.item_rows:
            self.item_rows[0]['qty_var'].set("")
            self.item_rows[0]['price_var'].set("")
            self.item_rows[0]['subtotal_label'].config(text="¥0.00")
            self.item_rows[0]['qty_entry'].focus()
        
        # 更新汇总
        self.update_summary()
    
    def refresh_display(self):
        """刷新显示"""
        if self.showing_today_only:
            self.update_tree_today()
        else:
            self.update_tree_all()
        self.update_stats()
    
    def update_tree_all(self):
        """显示所有记录（树形结构：销售为父节点，退货为子节点）"""
        # 清空现有数据
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 构建父子关系映射：original_record_id -> [退货记录列表]
        returns_by_parent = {}
        sale_records = []
        orphan_returns = []  # 没有关联原记录的退货
        
        for record in self.records:
            if record.get('type') == 'return' or record['quantity'] < 0:
                parent_id = record.get('original_record_id')
                if parent_id:
                    if parent_id not in returns_by_parent:
                        returns_by_parent[parent_id] = []
                    returns_by_parent[parent_id].append(record)
                else:
                    orphan_returns.append(record)
            else:
                sale_records.append(record)
        
        # 按日期排序（降序）
        sale_records = sorted(sale_records, key=lambda x: x['date'], reverse=True)
        orphan_returns = sorted(orphan_returns, key=lambda x: x['date'], reverse=True)
        
        # 插入销售记录（父节点）及其退货（子节点）
        for record in sale_records:
            parent_iid = self._insert_record(record, parent='')
            # 插入关联的退货记录作为子节点
            child_returns = returns_by_parent.get(record['id'], [])
            for ret_record in sorted(child_returns, key=lambda x: x.get('created_at', '')):
                self._insert_record(ret_record, parent=parent_iid)
        
        # 插入孤立的退货记录（没有关联原记录的）
        for record in orphan_returns:
            self._insert_record(record, parent='')

        # 设置退货记录的颜色
        self.tree.tag_configure('return', foreground='#e74c3c', background='#fef9f9')
        self.tree.tag_configure('child_return', foreground='#e74c3c', background='#fef9f9')
        self.update_total()

    def update_tree_today(self):
        """只显示今天记录（树形结构）"""
        # 清空现有数据
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        today = datetime.now().strftime("%Y-%m-%d")
        today_records = [r for r in self.records if r['date'] == today]
        
        # 构建父子关系映射
        returns_by_parent = {}
        sale_records = []
        orphan_returns = []
        
        for record in today_records:
            if record.get('type') == 'return' or record['quantity'] < 0:
                parent_id = record.get('original_record_id')
                if parent_id:
                    if parent_id not in returns_by_parent:
                        returns_by_parent[parent_id] = []
                    returns_by_parent[parent_id].append(record)
                else:
                    orphan_returns.append(record)
            else:
                sale_records.append(record)
        
        # 按创建时间排序（降序）
        sale_records = sorted(sale_records, key=lambda x: x.get('created_at', ''), reverse=True)
        orphan_returns = sorted(orphan_returns, key=lambda x: x.get('created_at', ''), reverse=True)
        
        # 插入销售记录（父节点）及其退货（子节点）
        for record in sale_records:
            parent_iid = self._insert_record(record, parent='')
            child_returns = returns_by_parent.get(record['id'], [])
            for ret_record in sorted(child_returns, key=lambda x: x.get('created_at', '')):
                self._insert_record(ret_record, parent=parent_iid)
        
        # 插入孤立的退货记录
        for record in orphan_returns:
            self._insert_record(record, parent='')

        # 设置退货记录的颜色
        self.tree.tag_configure('return', foreground='#e74c3c', background='#fef9f9')
        self.tree.tag_configure('child_return', foreground='#e74c3c', background='#fef9f9')
        self.update_total()

    def _display_records_tree(self, records_list):
        """通用树形显示方法：将记录按父子关系显示"""
        # 清空现有数据
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 构建父子关系映射
        returns_by_parent = {}
        sale_records = []
        orphan_returns = []
        
        for record in records_list:
            if record.get('type') == 'return' or record['quantity'] < 0:
                parent_id = record.get('original_record_id')
                if parent_id:
                    if parent_id not in returns_by_parent:
                        returns_by_parent[parent_id] = []
                    returns_by_parent[parent_id].append(record)
                else:
                    orphan_returns.append(record)
            else:
                sale_records.append(record)
        
        # 按日期排序（降序）
        sale_records = sorted(sale_records, key=lambda x: x['date'], reverse=True)
        orphan_returns = sorted(orphan_returns, key=lambda x: x['date'], reverse=True)
        
        # 插入销售记录（父节点）及其退货（子节点）
        for record in sale_records:
            parent_iid = self._insert_record(record, parent='')
            child_returns = returns_by_parent.get(record['id'], [])
            for ret_record in sorted(child_returns, key=lambda x: x.get('created_at', '')):
                self._insert_record(ret_record, parent=parent_iid)
        
        # 插入孤立的退货记录
        for record in orphan_returns:
            self._insert_record(record, parent='')

        # 设置退货记录的颜色
        self.tree.tag_configure('return', foreground='#e74c3c', background='#fef9f9')
        self.tree.tag_configure('child_return', foreground='#e74c3c', background='#fef9f9')
        self.update_total()

    def _insert_record(self, record, parent=''):
        """插入单条记录（兼容新旧数据格式，支持树形父子结构）"""
        quantity = record['quantity']
        total = record['total_amount']
        
        # 判断是否为退货
        is_return = record.get('type') == 'return' or quantity < 0
        is_child = parent != ''  # 是否为子节点（关联退货）
        
        # 获取备注
        note = record.get('note', '')
        
        # 格式化明细显示
        items = record.get('items', [])
        if items:
            # 多商品记录，生成明细字符串
            item_parts = []
            for item in items:
                qty = abs(item.get('quantity', 0))
                price = item.get('unit_price', 0)
                subtotal = qty * price
                item_parts.append(f"{qty}套@{price:.0f}={subtotal:.0f}")
            detail_display = ", ".join(item_parts)
            # 截断过长的明细
            if len(detail_display) > 25:
                detail_display = detail_display[:22] + "..."
        else:
            # 兼容旧数据：单商品
            qty = abs(quantity)
            price = record.get('unit_price', 0)
            detail_display = f"{qty}套@{price:.0f}={total:.0f}"
        
        # 格式化显示
        if is_return:
            qty_display = f"-{abs(quantity)}"
            total_display = f"-¥{abs(total):.2f}"
            note_display = note[:15] + ('...' if len(note) > 15 else '')
            if is_child:
                tags = ('child_return',)
            else:
                tags = ('return',)
        else:
            qty_display = str(quantity)
            total_display = f"¥{total:.2f}"
            note_display = note[:15] + ('...' if len(note) > 15 else '')
            tags = ()
        
        # 树形显示文本（子节点显示↳符号）
        tree_text = "↳" if is_child else ""
        
        item_iid = self.tree.insert(parent, tk.END, text=tree_text, values=(
            record['id'],
            record['date'],
            qty_display,
            detail_display,
            total_display,
            note_display
        ), tags=tags, open=True)
        
        return item_iid

    def update_total(self):
        """更新合计金额（包括树形子节点）"""
        if not hasattr(self, 'total_label'):
            return

        total_amount = 0.0

        def sum_children(parent):
            """递归计算所有节点金额"""
            nonlocal total_amount
            for item in self.tree.get_children(parent):
                item_values = self.tree.item(item, 'values')
                if item_values and len(item_values) >= 5:
                    amount_str = item_values[4]
                    amount_str = amount_str.replace('¥', '').replace(',', '').strip()
                    try:
                        amount = float(amount_str)
                        total_amount += amount
                    except:
                        pass
                # 递归处理子节点
                sum_children(item)
        
        sum_children('')

        # 格式化显示
        if abs(total_amount) < 0.01:
            self.total_label.config(text="¥0.00", fg='#2c3e50')
        elif total_amount < 0:
            self.total_label.config(text=f"-¥{abs(total_amount):.2f}", fg='#e74c3c')
        else:
            self.total_label.config(text=f"¥{total_amount:.2f}", fg='#27ae60')

    def update_tree(self, records=None):
        """更新表格（兼容旧方法）"""
        # 清空现有数据
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 显示记录
        if records is None:
            if self.showing_today_only:
                today = datetime.now().strftime("%Y-%m-%d")
                records = sorted(
                    [r for r in self.records if r['date'] == today],
                    key=lambda x: x.get('created_at', ''),
                    reverse=True
                )
            else:
                records = sorted(self.records, key=lambda x: x['date'], reverse=True)
        
        for record in records:
            quantity = record['quantity']
            total = record['total_amount']
            
            # 判断是否为退货
            is_return = record.get('type') == 'return' or quantity < 0
            
            # 格式化显示
            if is_return:
                qty_display = f"-{abs(quantity)}"
                total_display = f"-¥{abs(total):.2f}"
                note_display = record.get('note', '')[:18] + ('...' if len(record.get('note', '')) > 18 else '')
                tags = ('return',)
            else:
                qty_display = str(quantity)
                total_display = f"¥{total:.2f}"
                note_display = record.get('note', '')[:20] + ('...' if len(record.get('note', '')) > 20 else '')
                tags = ()
            
            item = self.tree.insert('', tk.END, values=(
                record['id'],
                record['date'],
                qty_display,
                f"¥{record['unit_price']:.2f}",
                total_display,
                note_display
            ), tags=tags)
        
        # 设置退货记录的颜色
        self.tree.tag_configure('return', foreground='#e74c3c', background='#fef9f9')
        self.update_total()

    def update_stats(self):
        """更新统计 - 深蓝风格"""
        today = datetime.now().strftime("%Y-%m-%d")
        today_records = [r for r in self.records if r['date'] == today]
        
        # 分离销售和退货
        sale_records = [r for r in today_records if r.get('type') != 'return' and r['quantity'] > 0]
        return_records = [r for r in today_records if r.get('type') == 'return' or r['quantity'] < 0]
        
        # 销售统计
        sale_qty = sum(r['quantity'] for r in sale_records)
        sale_amount = sum(r['total_amount'] for r in sale_records)
        
        # 退货统计
        return_qty = sum(abs(r['quantity']) for r in return_records)
        return_amount = sum(abs(r['total_amount']) for r in return_records)
        
        # 净统计
        net_qty = sale_qty - return_qty
        net_amount = sale_amount - return_amount
        
        # 深蓝风格统计
        stats_text = f"""📅 今日 {today}
━━━━━━━━━━━━━━
✅ 销售: {sale_qty}套 ¥{sale_amount:.0f}
🔄 退货: {return_qty}套 ¥{return_amount:.0f}
━━━━━━━━━━━━━━
📦 净额: {net_qty}套 ¥{net_amount:.0f}
"""
        self.stats_label.config(text=stats_text)

    def confirm_date_filter(self):
        """确认查看选中日期的记录（树形结构）"""
        year = self.filter_year_var.get()
        month = self.filter_month_var.get().zfill(2)
        day = self.filter_day_var.get().zfill(2)

        # 组合日期
        date_str = f"{year}-{month}-{day}"

        # 筛选该日期的记录
        filtered = [r for r in self.records if r['date'] == date_str]

        # 更新状态
        self.showing_today_only = False
        self.status_label.config(text=date_str, fg=self.COLORS['gray'])

        if filtered:
            self._display_records_tree(filtered)
            # 更新统计为选中日期的统计
            self.update_stats_for_date(date_str)
        else:
            # 清空表格
            for item in self.tree.get_children():
                self.tree.delete(item)
            # 没有记录时显示空状态
            self.stats_label.config(
                text=f"{date_str}\n暂无记录",
                justify=tk.CENTER
            )

    def update_stats_for_date(self, date_str):
        """更新指定日期的统计"""
        date_records = [r for r in self.records if r['date'] == date_str]

        if not date_records:
            return

        # 分离销售和退货
        sale_records = [r for r in date_records if r.get('type') != 'return' and r['quantity'] > 0]
        return_records = [r for r in date_records if r.get('type') == 'return' or r['quantity'] < 0]

        # 销售统计
        sale_qty = sum(r['quantity'] for r in sale_records)
        sale_amount = sum(r['total_amount'] for r in sale_records)

        # 退货统计
        return_qty = sum(abs(r['quantity']) for r in return_records)
        return_amount = sum(abs(r['total_amount']) for r in return_records)

        # 净统计
        net_qty = sale_qty - return_qty
        net_amount = sale_amount - return_amount

        avg_price = sale_amount / sale_qty if sale_qty > 0 else 0

        stats_text = f"""
📅 {date_str}
━━━━━━━━━━━━━━━━
✅ 销售: {sale_qty}套 ¥{sale_amount:.2f}
🔄 退货: {return_qty}套 ¥{return_amount:.2f}
━━━━━━━━━━━━━━━━
📦 净数量: {net_qty} 套
💵 净金额: ¥{net_amount:.2f}
💰 平均单价: ¥{avg_price:.2f}
📝 记录数: {len(date_records)} 条
        """
        self.stats_label.config(text=stats_text)

    def show_all_records(self):
        """显示所有记录"""
        self.showing_today_only = False
        self.update_tree_all()
        # 更新状态标签
        self.status_label.config(text="📋 全部", fg=self.COLORS['text_light'])

    def show_today_records(self):
        """显示今日记录"""
        self.showing_today_only = True
        self.update_tree_today()
        # 更新筛选框显示今天
        today = datetime.now()
        self.filter_year_var.set(str(today.year))
        self.filter_month_var.set(str(today.month).zfill(2))
        self.filter_day_var.set(str(today.day).zfill(2))
        # 更新状态标签
        self.status_label.config(text="🔴 今日", fg=self.COLORS['primary'])

    def show_month_records(self):
        """显示本月记录（树形结构）"""
        self.showing_today_only = False
        this_month = datetime.now().strftime("%Y-%m")
        month_records = [r for r in self.records if r['date'].startswith(this_month)]
        self._display_records_tree(month_records)
        # 更新筛选框和状态
        self.filter_year_var.set(str(datetime.now().year))
        self.filter_month_var.set(str(datetime.now().month).zfill(2))
        self.filter_day_var.set("01")
        self.status_label.config(text="📆 本月", fg=self.COLORS['text_light'])

    def show_year_records(self):
        """显示本年记录（树形结构）"""
        self.showing_today_only = False
        this_year = datetime.now().strftime("%Y")
        year_records = [r for r in self.records if r['date'].startswith(this_year)]
        self._display_records_tree(year_records)
        # 更新筛选框和状态
        self.filter_year_var.set(str(datetime.now().year))
        self.filter_month_var.set("01")
        self.filter_day_var.set("01")
        self.status_label.config(text="📊 本年", fg=self.COLORS['text_light'])

    def show_record_details(self, event=None):
        """显示选中记录的完整商品明细"""
        selected = self.tree.selection()
        if not selected:
            return
        
        item = selected[0]
        values = self.tree.item(item, 'values')
        if not values or len(values) < 5:
            return
        
        record_id = int(values[0])
        
        # 找到对应记录
        record = None
        for r in self.records:
            if r['id'] == record_id:
                record = r
                break
        
        if not record:
            return
        
        # 创建明细窗口
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"商品明细 - 记录#{record_id}")
        
        # 计算窗口位置：显示在主窗口中间
        self.root.update_idletasks()
        main_x = self.root.winfo_x()
        main_y = self.root.winfo_y()
        main_width = self.root.winfo_width()
        main_height = self.root.winfo_height()
        
        # 弹出窗口尺寸
        popup_width = 500
        popup_height = 450
        
        # 位置：主窗口中间
        popup_x = main_x + (main_width - popup_width) // 2
        popup_y = main_y + (main_height - popup_height) // 2
        
        detail_window.geometry(f"{popup_width}x{popup_height}+{popup_x}+{popup_y}")
        detail_window.resizable(True, True)
        detail_window.transient(self.root)
        
        # 顶部信息区
        info_frame = tk.Frame(detail_window, bg=self.COLORS['light'])
        info_frame.pack(fill=tk.X, padx=10, pady=10)
        
        is_return = record.get('type') == 'return' or record['quantity'] < 0
        type_text = "退货" if is_return else "销售"
        type_color = self.COLORS['danger'] if is_return else self.COLORS['success']
        
        tk.Label(info_frame, text=f"日期: {record['date']}", 
                font=('微软雅黑', 11), bg=self.COLORS['light']).pack(anchor='w', pady=2)
        tk.Label(info_frame, text=f"类型: {type_text}", 
                font=('微软雅黑', 11, 'bold'), bg=self.COLORS['light'], fg=type_color).pack(anchor='w', pady=2)
        
        note_text = record.get('note', '')
        if note_text:
            tk.Label(info_frame, text=f"备注: {note_text}", 
                    font=('微软雅黑', 10), bg=self.COLORS['light'], fg=self.COLORS['gray']).pack(anchor='w', pady=2)
        
        # 明细表格区
        table_frame = tk.Frame(detail_window, bg=self.COLORS['white'], 
                              relief='solid', borderwidth=1)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 表头
        header_frame = tk.Frame(table_frame, bg=self.COLORS['primary'])
        header_frame.pack(fill=tk.X)
        
        tk.Label(header_frame, text="序号", font=('微软雅黑', 10, 'bold'),
                bg=self.COLORS['primary'], fg=self.COLORS['white'], width=5).pack(side=tk.LEFT, padx=3, pady=8)
        tk.Label(header_frame, text="类型", font=('微软雅黑', 10, 'bold'),
                bg=self.COLORS['primary'], fg=self.COLORS['white'], width=6).pack(side=tk.LEFT, padx=3, pady=8)
        tk.Label(header_frame, text="数量", font=('微软雅黑', 10, 'bold'),
                bg=self.COLORS['primary'], fg=self.COLORS['white'], width=8).pack(side=tk.LEFT, padx=3, pady=8)
        tk.Label(header_frame, text="单价", font=('微软雅黑', 10, 'bold'),
                bg=self.COLORS['primary'], fg=self.COLORS['white'], width=10).pack(side=tk.LEFT, padx=3, pady=8)
        tk.Label(header_frame, text="小计", font=('微软雅黑', 10, 'bold'),
                bg=self.COLORS['primary'], fg=self.COLORS['white'], width=10).pack(side=tk.LEFT, padx=3, pady=8)
        
        # 商品明细行
        items = record.get('items', [])
        if not items:
            items = [{
                'quantity': abs(record['quantity']),
                'unit_price': record.get('unit_price', 0)
            }]
        
        total_qty = 0
        total_return_qty = 0
        total_amount = 0.0
        total_return_amount = 0.0
        
        for i, item_data in enumerate(items):
            qty = item_data.get('quantity', 0)
            price = item_data.get('unit_price', 0)
            subtotal = qty * price
            
            # 判断是销售还是退货
            is_item_return = qty < 0
            item_type = "退货" if is_item_return else "销售"
            item_type_color = self.COLORS['danger'] if is_item_return else self.COLORS['success']
            
            if is_item_return:
                total_return_qty += abs(qty)
                total_return_amount += abs(subtotal)
            else:
                total_qty += qty
                total_amount += subtotal
            
            row_bg = self.COLORS['light'] if i % 2 == 0 else self.COLORS['white']
            
            row_frame = tk.Frame(table_frame, bg=row_bg)
            row_frame.pack(fill=tk.X)
            
            tk.Label(row_frame, text=str(i + 1), font=('微软雅黑', 10),
                    bg=row_bg, width=5).pack(side=tk.LEFT, padx=3, pady=6)
            tk.Label(row_frame, text=item_type, font=('微软雅黑', 10, 'bold'),
                    bg=row_bg, fg=item_type_color, width=6).pack(side=tk.LEFT, padx=3, pady=6)
            tk.Label(row_frame, text=f"{abs(qty)}套", font=('微软雅黑', 10),
                    bg=row_bg, width=8).pack(side=tk.LEFT, padx=3, pady=6)
            tk.Label(row_frame, text=f"¥{price:.2f}", font=('微软雅黑', 10),
                    bg=row_bg, width=10).pack(side=tk.LEFT, padx=3, pady=6)
            
            subtotal_text = f"-¥{abs(subtotal):.2f}" if is_item_return else f"¥{subtotal:.2f}"
            subtotal_color = self.COLORS['danger'] if is_item_return else self.COLORS['primary']
            tk.Label(row_frame, text=subtotal_text, font=('微软雅黑', 10, 'bold'),
                    bg=row_bg, fg=subtotal_color, width=10).pack(side=tk.LEFT, padx=3, pady=6)
        
        # 合计行
        summary_frame = tk.Frame(table_frame, bg=self.COLORS['primary_light'])
        summary_frame.pack(fill=tk.X)
        
        net_qty = total_qty - total_return_qty
        net_amount = total_amount - total_return_amount
        
        tk.Label(summary_frame, text="合计", font=('微软雅黑', 11, 'bold'),
                bg=self.COLORS['primary_light'], fg=self.COLORS['white'], width=5).pack(side=tk.LEFT, padx=3, pady=10)
        tk.Label(summary_frame, text="", bg=self.COLORS['primary_light'], width=6).pack(side=tk.LEFT, padx=3)
        tk.Label(summary_frame, text=f"{net_qty}套", font=('微软雅黑', 11, 'bold'),
                bg=self.COLORS['primary_light'], fg=self.COLORS['white'], width=8).pack(side=tk.LEFT, padx=3, pady=10)
        tk.Label(summary_frame, text="", bg=self.COLORS['primary_light'], width=10).pack(side=tk.LEFT, padx=3)
        
        amount_text = f"-¥{abs(net_amount):.2f}" if net_amount < 0 else f"¥{net_amount:.2f}"
        amount_color = '#ff6b6b' if net_amount < 0 else self.COLORS['white']
        tk.Label(summary_frame, text=amount_text, font=('微软雅黑', 12, 'bold'),
                bg=self.COLORS['primary_light'], fg=amount_color, width=10).pack(side=tk.LEFT, padx=3, pady=10)
        
        # 按钮区
        btn_frame = tk.Frame(detail_window, bg=self.COLORS['light'])
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 打印小票按钮
        def print_this_record():
            detail_window.destroy()
            self.print_selected_record()
        
        print_btn = tk.Button(btn_frame, text="打印小票", command=print_this_record,
                             font=('微软雅黑', 11), bg='#27ae60', fg='white',
                             width=12, relief='flat', cursor='hand2')
        print_btn.pack(side=tk.LEFT, padx=20)
        
        close_btn = tk.Button(btn_frame, text="关闭", command=detail_window.destroy,
                             font=('微软雅黑', 11), bg=self.COLORS['primary'], fg=self.COLORS['white'],
                             width=12, relief='flat', cursor='hand2')
        close_btn.pack(side=tk.RIGHT, padx=20)
        
        detail_window.bind('<Escape>', lambda e: detail_window.destroy())

    def show_context_menu(self, event):
        """显示右键菜单"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            # 使用 tk_popup 在指定位置显示菜单
            screen_width = self.root.winfo_screenwidth()
            menu_x = int(screen_width * 0.6)
            menu_y = event.y_root
            try:
                self.context_menu.tk_popup(menu_x, menu_y)
            finally:
                self.context_menu.grab_release()
    
    def edit_note(self):
        """编辑选中记录的备注"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要编辑的记录")
            return
        
        item = selected[0]
        values = self.tree.item(item, 'values')
        record_id = int(values[0])
        
        # 找到对应记录
        record = None
        for r in self.records:
            if r['id'] == record_id:
                record = r
                break
        
        if not record:
            messagebox.showerror("错误", "未找到记录")
            return
        
        # 创建编辑窗口
        edit_window = tk.Toplevel(self.root)
        edit_window.title(f"编辑备注 - 记录#{record_id}")
        
        # 计算窗口位置：显示在主窗口中间
        self.root.update_idletasks()
        main_x = self.root.winfo_x()
        main_y = self.root.winfo_y()
        main_width = self.root.winfo_width()
        main_height = self.root.winfo_height()
        
        popup_width = 400
        popup_height = 300
        popup_x = main_x + (main_width - popup_width) // 2
        popup_y = main_y + (main_height - popup_height) // 2
        
        edit_window.geometry(f"{popup_width}x{popup_height}+{popup_x}+{popup_y}")
        edit_window.transient(self.root)
        edit_window.grab_set()  # 模态窗口
        
        # 显示记录信息
        info_frame = tk.Frame(edit_window)
        info_frame.pack(fill=tk.X, padx=15, pady=10)
        
        tk.Label(info_frame, text=f"📅 日期: {record['date']}", font=('微软雅黑', 10)).pack(anchor='w')
        tk.Label(info_frame, text=f"📦 数量: {abs(record['quantity'])}套", font=('微软雅黑', 10)).pack(anchor='w')
        tk.Label(info_frame, text=f"💵 金额: ¥{abs(record['total_amount']):.2f}", font=('微软雅黑', 10)).pack(anchor='w')
        
        # 备注编辑区
        tk.Label(edit_window, text="📝 备注:", font=('微软雅黑', 11)).pack(anchor='w', padx=15, pady=(10, 5))
        
        note_text = tk.Text(edit_window, font=('微软雅黑', 10), width=40, height=4)
        note_text.pack(padx=15, fill=tk.X)
        note_text.insert('1.0', record.get('note', ''))
        note_text.focus_set()
        
        # 按钮区
        btn_frame = tk.Frame(edit_window)
        btn_frame.pack(pady=20)
        
        def save_note():
            new_note = note_text.get('1.0', tk.END).strip()
            record['note'] = new_note
            self.save_records()
            self.refresh_display()
            edit_window.destroy()
            messagebox.showinfo("成功", "备注已更新")
        
        tk.Button(btn_frame, text="✅ 确认", command=save_note,
                  font=('微软雅黑', 11), bg='#27ae60', fg='white', width=10, height=1).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="❌ 取消", command=edit_window.destroy,
                  font=('微软雅黑', 11), bg='#e74c3c', fg='white', width=10, height=1).pack(side=tk.LEFT, padx=10)
        
        # 绑定Ctrl+Enter保存
        note_text.bind('<Control-Return>', lambda e: save_note())
    
    def edit_quantity_price(self):
        """编辑选中记录的明细（支持添加/删除商品）"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要编辑的记录")
            return
        
        item = selected[0]
        values = self.tree.item(item, 'values')
        record_id = int(values[0])
        
        # 找到对应记录
        record = None
        for r in self.records:
            if r['id'] == record_id:
                record = r
                break
        
        if not record:
            messagebox.showerror("错误", "未找到记录")
            return
        
        # 创建编辑窗口
        edit_window = tk.Toplevel(self.root)
        edit_window.title(f"编辑明细 - 记录#{record_id}")
        
        # 计算窗口位置：显示在主窗口中间
        self.root.update_idletasks()
        main_x = self.root.winfo_x()
        main_y = self.root.winfo_y()
        main_width = self.root.winfo_width()
        main_height = self.root.winfo_height()
        
        popup_width = 420
        popup_height = 600
        popup_x = main_x + (main_width - popup_width) // 2
        popup_y = main_y + (main_height - popup_height) // 2
        
        edit_window.geometry(f"{popup_width}x{popup_height}+{popup_x}+{popup_y}")
        edit_window.resizable(True, True)
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # 显示记录信息
        info_frame = tk.Frame(edit_window)
        info_frame.pack(fill=tk.X, padx=15, pady=10)
        
        tk.Label(info_frame, text=f"📅 日期: {record['date']}", font=('微软雅黑', 10)).pack(anchor='w')
        tk.Label(info_frame, text=f"📝 备注: {record.get('note', '')[:30]}", font=('微软雅黑', 10)).pack(anchor='w')
        
        # 商品明细列表
        list_frame = tk.LabelFrame(edit_window, text="商品明细", font=('微软雅黑', 11))
        list_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
        
        # 商品列表（Listbox）
        listbox_frame = tk.Frame(list_frame)
        listbox_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        scrollbar = tk.Scrollbar(listbox_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        items_listbox = tk.Listbox(
            listbox_frame,
            font=('微软雅黑', 10),
            height=6,
            yscrollcommand=scrollbar.set
        )
        items_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=items_listbox.yview)
        
        # 当前商品列表
        items = record.get('items', [])
        if not items:
            # 兼容旧数据，创建单商品
            items = [{
                'quantity': abs(record['quantity']),
                'unit_price': record['unit_price']
            }]
        record['items'] = items
        
        # 先创建Label（后面refresh_list会用到）
        total_qty_label = tk.Label(edit_window, text="总数量: 0套", font=('微软雅黑', 10))
        total_amount_label = tk.Label(edit_window, text="总金额: ¥0.00", font=('微软雅黑', 10))
        
        # 填充列表
        def refresh_list():
            items_listbox.delete(0, tk.END)
            total_qty = 0
            total_amount = 0.0
            for i, item in enumerate(items):
                qty = item.get('quantity', 0)
                price = item.get('unit_price', 0)
                subtotal = qty * price
                total_qty += qty
                total_amount += subtotal
                items_listbox.insert(tk.END, f"{qty}套 @ ¥{price:.0f} = ¥{subtotal:.0f}    [删除]")
            total_qty_label.config(text=f"总数量: {total_qty}套")
            total_amount_label.config(text=f"总金额: ¥{total_amount:.2f}")
            return total_qty, total_amount
        
        current_total_qty, current_total_amount = refresh_list()
        
        def delete_item():
            """删除选中商品"""
            selected_idx = items_listbox.curselection()
            if not selected_idx:
                messagebox.showwarning("提示", "请先选择要删除的商品")
                return
            
            idx = selected_idx[0]
            del items[idx]
            refresh_list()
        
        # 操作按钮区（删除+添加）
        action_frame = tk.Frame(edit_window)
        action_frame.pack(fill=tk.X, padx=15, pady=10)
        
        del_btn = tk.Button(action_frame, text="🗑️ 删除选中", command=delete_item,
                          font=('微软雅黑', 10), bg='#e74c3c', fg='white', width=12)
        del_btn.pack(side=tk.LEFT, padx=5)
        
        def show_add_dialog():
            """弹出添加商品对话框（支持多行）"""
            dialog = tk.Toplevel(edit_window)
            dialog.title("添加商品")
            
            # 计算窗口位置：显示在编辑明细窗口右边
            edit_window.update_idletasks()
            edit_x = edit_window.winfo_x()
            edit_y = edit_window.winfo_y()
            edit_width = edit_window.winfo_width()
            edit_height = edit_window.winfo_height()
            
            popup_width = 400
            popup_height = 400
            popup_x = edit_x + edit_width + 10  # 编辑明细窗口右边
            popup_y = edit_y + (edit_height - popup_height) // 2  # 垂直居中
            
            dialog.geometry(f"{popup_width}x{popup_height}+{popup_x}+{popup_y}")
            dialog.transient(edit_window)
            dialog.grab_set()
            
            # 存储所有商品行的数据
            new_items = []
            item_rows = []
            
            # 商品列表区域
            list_frame = tk.LabelFrame(dialog, text="待添加商品", font=('微软雅黑', 10))
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            # 表头
            header_frame = tk.Frame(list_frame)
            header_frame.pack(fill=tk.X, padx=5, pady=2)
            tk.Label(header_frame, text="数量", font=('微软雅黑', 9), width=8).pack(side=tk.LEFT)
            tk.Label(header_frame, text="单价", font=('微软雅黑', 9), width=8).pack(side=tk.LEFT)
            tk.Label(header_frame, text="小计", font=('微软雅黑', 9), width=10).pack(side=tk.LEFT)
            
            # 商品行容器
            rows_container = tk.Frame(list_frame)
            rows_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            def add_item_row():
                """添加一行商品输入"""
                row_frame = tk.Frame(rows_container)
                row_frame.pack(fill=tk.X, pady=2)
                
                qty_var = tk.StringVar()
                price_var = tk.StringVar()
                
                qty_entry = tk.Entry(row_frame, textvariable=qty_var, font=('微软雅黑', 10), width=8)
                qty_entry.pack(side=tk.LEFT, padx=2)
                
                price_entry = tk.Entry(row_frame, textvariable=price_var, font=('微软雅黑', 10), width=8)
                price_entry.pack(side=tk.LEFT, padx=2)
                
                subtotal_label = tk.Label(row_frame, text="¥0.00", font=('微软雅黑', 10), width=10, anchor='w')
                subtotal_label.pack(side=tk.LEFT, padx=2)
                
                def update_subtotal(*args):
                    try:
                        qty = int(qty_var.get() or 0)
                        price = float(price_var.get() or 0)
                        subtotal_label.config(text=f"¥{qty * price:.2f}")
                    except:
                        subtotal_label.config(text="¥0.00")
                
                qty_var.trace_add('write', update_subtotal)
                price_var.trace_add('write', update_subtotal)
                
                def delete_row():
                    if len(item_rows) > 1:
                        row_frame.destroy()
                        item_rows.remove(row_data)
                
                del_btn = tk.Button(row_frame, text="🗑", command=delete_row, 
                                   font=('微软雅黑', 8), bg='#e74c3c', fg='white', width=2)
                del_btn.pack(side=tk.LEFT, padx=2)
                
                row_data = {
                    'qty_var': qty_var,
                    'price_var': price_var,
                    'frame': row_frame
                }
                item_rows.append(row_data)
                
                # 回车跳转
                qty_entry.bind('<Return>', lambda e: price_entry.focus())
                price_entry.bind('<Return>', lambda e: add_item_row())
                
                qty_entry.focus_set()
                return row_data
            
            # 添加第一行
            add_item_row()
            
            # 添加商品行按钮
            add_row_btn = tk.Button(dialog, text="➕ 添加商品行", command=add_item_row,
                                   font=('微软雅黑', 10), bg='#3498db', fg='white')
            add_row_btn.pack(pady=5)
            
            def do_add():
                """确认添加所有商品"""
                try:
                    added_count = 0
                    for row in item_rows:
                        qty_str = row['qty_var'].get().strip()
                        price_str = row['price_var'].get().strip()
                        
                        if qty_str and price_str:
                            qty = int(qty_str)
                            price = float(price_str)
                            if qty > 0 and price > 0:
                                items.append({'quantity': qty, 'unit_price': price})
                                added_count += 1
                    
                    if added_count > 0:
                        refresh_list()
                        dialog.destroy()
                        messagebox.showinfo("成功", f"已添加 {added_count} 个商品")
                    else:
                        messagebox.showwarning("提示", "请至少填写一个有效的商品")
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的数字")
            
            # 按钮区
            btn_frame = tk.Frame(dialog)
            btn_frame.pack(pady=10)
            
            tk.Button(btn_frame, text="✅ 确认添加", command=do_add,
                     font=('微软雅黑', 11), bg='#27ae60', fg='white', width=12).pack(side=tk.LEFT, padx=10)
            tk.Button(btn_frame, text="❌ 取消", command=dialog.destroy,
                     font=('微软雅黑', 11), bg='#e74c3c', fg='white', width=12).pack(side=tk.LEFT, padx=10)
            
            # 快捷键：Ctrl+Enter 确认添加
            def on_ctrl_enter(event):
                do_add()
                return "break"
            dialog.bind('<Control-Return>', on_ctrl_enter)
        
        add_btn = tk.Button(action_frame, text="➕ 添加商品", command=show_add_dialog,
                          font=('微软雅黑', 10), bg='#3498db', fg='white', width=12)
        add_btn.pack(side=tk.LEFT, padx=20)
        
        # 添加退货按钮
        def show_add_return_dialog():
            """显示添加退货对话框"""
            dialog = tk.Toplevel(edit_window)
            dialog.title("添加退货")
            
            # 计算窗口位置：显示在编辑明细窗口右边
            edit_window.update_idletasks()
            edit_x = edit_window.winfo_x()
            edit_y = edit_window.winfo_y()
            edit_width = edit_window.winfo_width()
            edit_height = edit_window.winfo_height()
            
            popup_width = 400
            popup_height = 400
            popup_x = edit_x + edit_width + 10  # 编辑明细窗口右边
            popup_y = edit_y + (edit_height - popup_height) // 2  # 垂直居中
            
            dialog.geometry(f"{popup_width}x{popup_height}+{popup_x}+{popup_y}")
            dialog.transient(edit_window)
            dialog.grab_set()
            
            # 存储所有商品行的数据
            item_rows = []
            
            # 商品列表区域
            list_frame = tk.LabelFrame(dialog, text="退货商品", font=('微软雅黑', 10))
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            # 表头
            header_frame = tk.Frame(list_frame)
            header_frame.pack(fill=tk.X, padx=5, pady=2)
            tk.Label(header_frame, text="数量", font=('微软雅黑', 9), width=8).pack(side=tk.LEFT)
            tk.Label(header_frame, text="单价", font=('微软雅黑', 9), width=8).pack(side=tk.LEFT)
            tk.Label(header_frame, text="小计", font=('微软雅黑', 9), width=10).pack(side=tk.LEFT)
            
            # 商品行容器
            rows_container = tk.Frame(list_frame)
            rows_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            def add_item_row():
                """添加一行商品输入"""
                row_frame = tk.Frame(rows_container)
                row_frame.pack(fill=tk.X, pady=2)
                
                qty_var = tk.StringVar()
                price_var = tk.StringVar()
                
                qty_entry = tk.Entry(row_frame, textvariable=qty_var, font=('微软雅黑', 10), width=8)
                qty_entry.pack(side=tk.LEFT, padx=2)
                
                price_entry = tk.Entry(row_frame, textvariable=price_var, font=('微软雅黑', 10), width=8)
                price_entry.pack(side=tk.LEFT, padx=2)
                
                subtotal_label = tk.Label(row_frame, text="-¥0.00", font=('微软雅黑', 10), width=10, 
                                         anchor='w', fg='#e74c3c')
                subtotal_label.pack(side=tk.LEFT, padx=2)
                
                def update_subtotal(*args):
                    try:
                        qty = int(qty_var.get() or 0)
                        price = float(price_var.get() or 0)
                        subtotal_label.config(text=f"-¥{qty * price:.2f}")
                    except:
                        subtotal_label.config(text="-¥0.00")
                
                qty_var.trace_add('write', update_subtotal)
                price_var.trace_add('write', update_subtotal)
                
                def delete_row():
                    if len(item_rows) > 1:
                        row_frame.destroy()
                        item_rows.remove(row_data)
                
                del_btn = tk.Button(row_frame, text="🗑", command=delete_row, 
                                   font=('微软雅黑', 8), bg='#e74c3c', fg='white', width=2)
                del_btn.pack(side=tk.LEFT, padx=2)
                
                row_data = {
                    'qty_var': qty_var,
                    'price_var': price_var,
                    'frame': row_frame
                }
                item_rows.append(row_data)
                
                qty_entry.bind('<Return>', lambda e: price_entry.focus())
                price_entry.bind('<Return>', lambda e: add_item_row())
                
                qty_entry.focus_set()
                return row_data
            
            # 添加第一行
            add_item_row()
            
            # 添加商品行按钮
            add_row_btn = tk.Button(dialog, text="➕ 添加退货行", command=add_item_row,
                                   font=('微软雅黑', 10), bg='#e74c3c', fg='white')
            add_row_btn.pack(pady=5)
            
            def do_add():
                """确认添加退货商品"""
                try:
                    added_count = 0
                    for row in item_rows:
                        qty_str = row['qty_var'].get().strip()
                        price_str = row['price_var'].get().strip()
                        
                        if qty_str and price_str:
                            qty = int(qty_str)
                            price = float(price_str)
                            if qty > 0 and price > 0:
                                # 退货商品数量为负
                                items.append({'quantity': -qty, 'unit_price': price})
                                added_count += 1
                    
                    if added_count > 0:
                        refresh_list()
                        dialog.destroy()
                        messagebox.showinfo("成功", f"已添加 {added_count} 个退货商品")
                    else:
                        messagebox.showwarning("提示", "请至少填写一个有效的商品")
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的数字")
            
            # 按钮区
            btn_frame = tk.Frame(dialog)
            btn_frame.pack(pady=10)
            
            tk.Button(btn_frame, text="✅ 确认添加", command=do_add,
                     font=('微软雅黑', 11), bg='#e74c3c', fg='white', width=12).pack(side=tk.LEFT, padx=10)
            tk.Button(btn_frame, text="❌ 取消", command=dialog.destroy,
                     font=('微软雅黑', 11), bg='#95a5a6', fg='white', width=12).pack(side=tk.LEFT, padx=10)
            
            def on_ctrl_enter(event):
                do_add()
                return "break"
            dialog.bind('<Control-Return>', on_ctrl_enter)
        
        add_return_btn = tk.Button(action_frame, text="➖ 添加退货", command=show_add_return_dialog,
                                  font=('微软雅黑', 10), bg='#e74c3c', fg='white', width=12)
        add_return_btn.pack(side=tk.LEFT, padx=20)
        
        # 总计区
        total_frame = tk.Frame(edit_window, bg='#ecf0f1')
        total_frame.pack(fill=tk.X, padx=15, pady=10)
        
        # 把Label放到总计区（不再重新定义变量）
        total_qty_label.config(font=('微软雅黑', 11, 'bold'), bg='#ecf0f1')
        total_qty_label.pack(in_=total_frame, side=tk.LEFT, padx=20)
        
        total_amount_label.config(font=('微软雅黑', 11, 'bold'), bg='#ecf0f1', fg='#e74c3c')
        total_amount_label.pack(in_=total_frame, side=tk.RIGHT, padx=20)
        
        # 按钮区
        btn_frame = tk.Frame(edit_window)
        btn_frame.pack(pady=15)
        
        def save_changes():
            """保存修改"""
            if not items:
                messagebox.showerror("错误", "请至少添加一个商品")
                return
            
            # 保持原有的正负号（退货记录数量为负）
            is_return = record.get('type') == 'return' or record['quantity'] < 0
            
            # 计算总数量和总金额
            total_qty = sum(item.get('quantity', 0) for item in items)
            total_amount = sum(item.get('quantity', 0) * item.get('unit_price', 0) for item in items)
            
            if is_return:
                record['quantity'] = -total_qty
                record['total_amount'] = -total_amount
            else:
                record['quantity'] = total_qty
                record['total_amount'] = total_amount
            
            # 保持unit_price为第一个商品的单价（兼容旧数据）
            record['unit_price'] = items[0].get('unit_price', 0)
            record['items'] = items
            
            self.save_records()
            self.refresh_display()
            edit_window.destroy()
            messagebox.showinfo("成功", "记录已更新")
        
        tk.Button(btn_frame, text="✅ 确认", command=save_changes,
                  font=('微软雅黑', 11), bg='#27ae60', fg='white', width=10).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="❌ 取消", command=edit_window.destroy,
                  font=('微软雅黑', 11), bg='#e74c3c', fg='white', width=10).pack(side=tk.LEFT, padx=10)
        
        # 快捷键
        edit_window.bind('<Control-Return>', lambda e: save_changes())
    
    def convert_to_return(self):
        """部分退货 - 支持多个不同单价的商品"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要退货的记录")
            return
        
        item = selected[0]
        values = self.tree.item(item, 'values')
        record_id = int(values[0])
        
        # 找到对应记录
        record = None
        for r in self.records:
            if r['id'] == record_id:
                record = r
                break
        
        if not record:
            messagebox.showerror("错误", "未找到记录")
            return
        
        # 检查是否已经是退货记录
        if record.get('type') == 'return' or record['quantity'] < 0:
            messagebox.showinfo("提示", "该记录已经是退货记录，无法再退货")
            return
        
        # 创建退货窗口
        return_window = tk.Toplevel(self.root)
        return_window.title(f"部分退货 - 记录#{record_id}")
        
        # 计算窗口位置：显示在主窗口中间
        self.root.update_idletasks()
        main_x = self.root.winfo_x()
        main_y = self.root.winfo_y()
        main_width = self.root.winfo_width()
        main_height = self.root.winfo_height()
        
        popup_width = 450
        popup_height = 420
        popup_x = main_x + (main_width - popup_width) // 2
        popup_y = main_y + (main_height - popup_height) // 2
        
        return_window.geometry(f"{popup_width}x{popup_height}+{popup_x}+{popup_y}")
        return_window.transient(self.root)
        return_window.grab_set()
        
        # 显示原记录信息
        info_frame = tk.LabelFrame(return_window, text="原销售记录", font=('微软雅黑', 10, 'bold'))
        info_frame.pack(fill=tk.X, padx=15, pady=10)
        
        tk.Label(info_frame, text=f"📅 日期: {record['date']}  |  📦 数量: {record['quantity']}套  |  💵 金额: ¥{record['total_amount']:.2f}", 
                font=('微软雅黑', 10)).pack(anchor='w', padx=10, pady=5)
        
        # 退货商品明细区
        return_frame = tk.LabelFrame(return_window, text="退货商品明细（可添加多行）", font=('微软雅黑', 10, 'bold'))
        return_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        # 表头
        header_frame = tk.Frame(return_frame, bg='#ecf0f1')
        header_frame.pack(fill=tk.X, padx=5, pady=2)
        tk.Label(header_frame, text="数量", font=('微软雅黑', 9, 'bold'), bg='#ecf0f1', width=8).pack(side=tk.LEFT, padx=2)
        tk.Label(header_frame, text="单价", font=('微软雅黑', 9, 'bold'), bg='#ecf0f1', width=8).pack(side=tk.LEFT, padx=2)
        tk.Label(header_frame, text="小计", font=('微软雅黑', 9, 'bold'), bg='#ecf0f1', width=10).pack(side=tk.LEFT, padx=2)
        tk.Label(header_frame, text="", bg='#ecf0f1', width=3).pack(side=tk.LEFT, padx=2)
        
        # 商品行容器
        items_container = tk.Frame(return_frame)
        items_container.pack(fill=tk.X, padx=5, pady=2)
        
        return_items = []  # 存储退货商品行
        
        # 汇总显示
        summary_frame = tk.Frame(return_frame)
        summary_frame.pack(fill=tk.X, padx=5, pady=5)
        summary_label = tk.Label(summary_frame, text="退货汇总: 0套  ¥0.00", font=('微软雅黑', 11, 'bold'), fg='#e74c3c')
        summary_label.pack(anchor='w')
        
        def update_summary():
            total_qty = 0
            total_amount = 0.0
            for item_row in return_items:
                try:
                    qty = int(item_row['qty_var'].get() or 0)
                    price = float(item_row['price_var'].get() or 0)
                    subtotal = qty * price
                    item_row['subtotal_label'].config(text=f"¥{subtotal:.2f}")
                    total_qty += qty
                    total_amount += subtotal
                except:
                    item_row['subtotal_label'].config(text="¥0.00")
            summary_label.config(text=f"退货汇总: {total_qty}套  ¥{total_amount:.2f}")
        
        def add_return_row():
            row_frame = tk.Frame(items_container)
            row_frame.pack(fill=tk.X, pady=1)
            
            qty_var = tk.StringVar()
            price_var = tk.StringVar()
            
            qty_entry = tk.Entry(row_frame, textvariable=qty_var, font=('微软雅黑', 10), width=8)
            qty_entry.pack(side=tk.LEFT, padx=2)
            
            price_entry = tk.Entry(row_frame, textvariable=price_var, font=('微软雅黑', 10), width=8)
            price_entry.pack(side=tk.LEFT, padx=2)
            
            subtotal_label = tk.Label(row_frame, text="¥0.00", font=('微软雅黑', 10), width=10, anchor='w')
            subtotal_label.pack(side=tk.LEFT, padx=2)
            
            row_data = {
                'qty_var': qty_var,
                'price_var': price_var,
                'subtotal_label': subtotal_label,
                'frame': row_frame,
                'qty_entry': qty_entry,
                'price_entry': price_entry
            }
            
            def delete_row():
                if len(return_items) > 1:
                    row_frame.destroy()
                    return_items.remove(row_data)
                    update_summary()
            
            del_btn = tk.Button(row_frame, text="🗑", command=delete_row, font=('微软雅黑', 8), 
                               bg='#e74c3c', fg='white', width=2)
            del_btn.pack(side=tk.LEFT, padx=2)
            
            qty_var.trace_add('write', lambda *args: update_summary())
            price_var.trace_add('write', lambda *args: update_summary())
            
            return_items.append(row_data)
            
            qty_entry.bind('<Return>', lambda e: price_entry.focus())
            price_entry.bind('<Return>', lambda e: add_return_row() if qty_var.get() and price_var.get() else None)
            
            qty_entry.focus_set()
            return row_data
        
        # 添加第一行
        add_return_row()
        
        # 添加行按钮
        tk.Button(return_frame, text="➕ 添加退货商品", command=add_return_row,
                  font=('微软雅黑', 9), bg='#3498db', fg='white').pack(pady=5)
        
        # 按钮区
        btn_frame = tk.Frame(return_window)
        btn_frame.pack(pady=15)
        
        def do_return():
            # 收集所有退货商品
            items = []
            total_qty = 0
            total_amount = 0.0
            
            for item_row in return_items:
                try:
                    qty = int(item_row['qty_var'].get() or 0)
                    price = float(item_row['price_var'].get() or 0)
                    if qty > 0 and price > 0:
                        items.append({'quantity': -qty, 'unit_price': price})
                        total_qty += qty
                        total_amount += qty * price
                except:
                    pass
            
            if not items:
                messagebox.showerror("错误", "请至少添加一个有效的退货商品")
                return
            
            if total_qty > record['quantity']:
                messagebox.showerror("错误", f"退货总数量({total_qty}套)不能超过原购买数量({record['quantity']}套)")
                return
            
            # 计算平均单价
            avg_price = total_amount / total_qty if total_qty > 0 else 0
            
            # 创建退货记录
            return_record = {
                "id": len(self.records) + 1,
                "date": datetime.now().strftime("%Y-%m-%d"),
                "quantity": -total_qty,
                "unit_price": avg_price,
                "total_amount": -total_amount,
                "note": f"[退货] 原记录#{record_id} {record.get('note', '')}",
                "type": "return",
                "items": items,
                "original_record_id": record_id,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            self.records.append(return_record)
            self.save_records()
            self.refresh_display()
            return_window.destroy()

            # 显示成功提示，支持打印
            success_msg = f"退货成功！\n退货: {total_qty}套 ({len(items)}种)\n退款: ¥{total_amount:.2f}"
            self.show_success_message(success_msg, return_record)
        
        tk.Button(btn_frame, text="✅ 确认退货", command=do_return,
                  font=('微软雅黑', 11), bg='#e74c3c', fg='white', width=12).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="❌ 取消", command=return_window.destroy,
                  font=('微软雅黑', 11), bg='#95a5a6', fg='white', width=10).pack(side=tk.LEFT, padx=10)
        
        def on_ctrl_enter(event):
            do_return()
            return "break"
        return_window.bind('<Control-Return>', on_ctrl_enter)
    
    def delete_selected(self):
        """删除选中记录"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的记录")
            return
        
        if messagebox.askyesno("确认删除", "确定要删除这条记录吗？"):
            item = selected[0]
            values = self.tree.item(item, 'values')
            record_id = int(values[0])
            
            # 删除记录
            self.records = [r for r in self.records if r['id'] != record_id]
            
            # 重新编号
            for i, record in enumerate(self.records, 1):
                record['id'] = i
            
            self.save_records()
            self.refresh_display()
            messagebox.showinfo("成功", "记录已删除")
    
    def export_csv(self):
        """导出CSV"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"记账导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(['ID', '日期', '数量', '单价', '总金额', '备注', '创建时间'])
                    
                    for record in self.records:
                        writer.writerow([
                            record['id'],
                            record['date'],
                            record['quantity'],
                            record['unit_price'],
                            record['total_amount'],
                            record.get('note', ''),
                            record.get('created_at', '')
                        ])
                
                messagebox.showinfo("成功", f"数据已导出到:\n{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def import_csv(self):
        """导入CSV"""
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                imported = 0
                failed = 0
                
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    
                    for row in reader:
                        try:
                            # 自动识别列名
                            date = row.get('日期') or row.get('date') or row.get('Date')
                            quantity = row.get('数量') or row.get('quantity') or row.get('Quantity')
                            price = row.get('单价') or row.get('unit_price') or row.get('price')
                            note = row.get('备注') or row.get('note') or row.get('Note')
                            
                            if date and quantity and price:
                                record = {
                                    "id": len(self.records) + 1,
                                    "date": date,
                                    "quantity": int(float(quantity)),
                                    "unit_price": float(price),
                                    "total_amount": int(float(quantity)) * float(price),
                                    "note": note or "",
                                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                }
                                self.records.append(record)
                                imported += 1
                            else:
                                failed += 1
                        except:
                            failed += 1
                
                self.save_records()
                self.refresh_display()
                
                msg = f"导入完成！\n成功: {imported} 条"
                if failed > 0:
                    msg += f"\n失败: {failed} 条"
                messagebox.showinfo("导入结果", msg)
                
            except Exception as e:
                messagebox.showerror("错误", f"导入失败: {str(e)}")

    def import_excel(self):
        """导入Excel文件"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(
                "缺少依赖",
                "导入Excel需要安装 openpyxl 库\n\n"
                "请运行以下命令安装:\n"
                "pip install openpyxl"
            )
            return

        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("Excel 2007+", "*.xlsx"),
                ("Excel 97-2003", "*.xls"),
                ("All files", "*.*")
            ]
        )

        if not file_path:
            return

        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            # 获取表头
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else "")

            if not headers:
                messagebox.showerror("错误", "Excel文件没有表头行")
                return

            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("Excel导入预览")
            preview_window.geometry("800x600")
            preview_window.transient(self.root)

            # 说明文字
            tk.Label(
                preview_window,
                text="📊 请选择数据对应的列",
                font=('微软雅黑', 14, 'bold')
            ).pack(pady=10)

            # 列选择区域
            selection_frame = tk.Frame(preview_window)
            selection_frame.pack(pady=10)

            tk.Label(selection_frame, text="日期列:", font=('微软雅黑', 11)).grid(row=0, column=0, padx=5, pady=5)
            date_col = ttk.Combobox(selection_frame, values=headers, width=20, state='readonly')
            date_col.grid(row=0, column=1, padx=5, pady=5)

            tk.Label(selection_frame, text="数量列:", font=('微软雅黑', 11)).grid(row=1, column=0, padx=5, pady=5)
            qty_col = ttk.Combobox(selection_frame, values=headers, width=20, state='readonly')
            qty_col.grid(row=1, column=1, padx=5, pady=5)

            tk.Label(selection_frame, text="单价列:", font=('微软雅黑', 11)).grid(row=2, column=0, padx=5, pady=5)
            price_col = ttk.Combobox(selection_frame, values=headers, width=20, state='readonly')
            price_col.grid(row=2, column=1, padx=5, pady=5)

            tk.Label(selection_frame, text="备注列(可选):", font=('微软雅黑', 11)).grid(row=3, column=0, padx=5, pady=5)
            note_col = ttk.Combobox(selection_frame, values=["(无)"] + headers, width=20, state='readonly')
            note_col.set("(无)")
            note_col.grid(row=3, column=1, padx=5, pady=5)

            # 自动识别常用列名
            for i, h in enumerate(headers):
                h_lower = h.lower()
                if any(kw in h_lower for kw in ['日期', 'date', '时间']):
                    date_col.set(h)
                elif any(kw in h_lower for kw in ['数量', 'quantity', '套数', '件数', '套']):
                    qty_col.set(h)
                elif any(kw in h_lower for kw in ['单价', 'price', 'unit', '价格']):
                    price_col.set(h)
                elif any(kw in h_lower for kw in ['备注', 'note', '说明', '客户']):
                    note_col.set(h)

            # 预览表格
            preview_frame = tk.LabelFrame(preview_window, text="数据预览（前10行）", font=('微软雅黑', 11))
            preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            # 创建预览表格
            tree_frame = tk.Frame(preview_frame)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

            scrollbar_y = tk.Scrollbar(tree_frame)
            scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

            preview_tree = ttk.Treeview(
                tree_frame,
                columns=headers,
                show='headings',
                yscrollcommand=scrollbar_y.set,
                height=10
            )

            for h in headers:
                preview_tree.heading(h, text=h)
                preview_tree.column(h, width=100)

            preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar_y.config(command=preview_tree.yview)

            # 加载预览数据
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row_idx > 11:  # 只显示前10行
                    break
                values = [str(v) if v is not None else "" for v in row]
                preview_tree.insert('', tk.END, values=values)

            # 导入按钮
            def do_import():
                d_col = date_col.get()
                q_col = qty_col.get()
                p_col = price_col.get()
                n_col = note_col.get()

                if not d_col or not q_col or not p_col:
                    messagebox.showwarning("提示", "请选择日期、数量和单价列")
                    return

                # 获取列索引
                d_idx = headers.index(d_col)
                q_idx = headers.index(q_col)
                p_idx = headers.index(p_col)
                n_idx = headers.index(n_col) if n_col != "(无)" else None

                imported = 0
                failed = 0
                failed_rows = []

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        date_val = row[d_idx]
                        qty_val = row[q_idx]
                        price_val = row[p_idx]
                        note_val = row[n_idx] if n_idx is not None else ""

                        # 处理日期
                        if isinstance(date_val, datetime):
                            date_str = date_val.strftime("%Y-%m-%d")
                        elif isinstance(date_val, str):
                            # 尝试解析日期字符串
                            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y年%m月%d日"]:
                                try:
                                    date_str = datetime.strptime(date_val.strip(), fmt).strftime("%Y-%m-%d")
                                    break
                                except:
                                    continue
                            else:
                                failed += 1
                                failed_rows.append(f"第{row_idx}行: 日期格式无法识别")
                                continue
                        else:
                            # 尝试Excel序列号
                            try:
                                excel_date = int(float(date_val))
                                parsed = datetime(1899, 12, 30) + __import__('datetime').timedelta(days=excel_date)
                                date_str = parsed.strftime("%Y-%m-%d")
                            except:
                                failed += 1
                                failed_rows.append(f"第{row_idx}行: 日期格式无效")
                                continue

                        # 处理数量和单价
                        quantity = float(qty_val) if qty_val is not None else 0
                        price = float(price_val) if price_val is not None else 0

                        if quantity <= 0 or price <= 0:
                            failed += 1
                            failed_rows.append(f"第{row_idx}行: 数量或单价无效")
                            continue

                        # 处理备注
                        note = str(note_val) if note_val is not None else ""

                        # 创建记录
                        record = {
                            "id": len(self.records) + 1,
                            "date": date_str,
                            "quantity": int(quantity),
                            "unit_price": float(price),
                            "total_amount": int(quantity) * float(price),
                            "note": note,
                            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }

                        self.records.append(record)
                        imported += 1

                    except Exception as e:
                        failed += 1
                        failed_rows.append(f"第{row_idx}行: {str(e)}")

                self.save_records()
                self.refresh_display()

                preview_window.destroy()

                # 显示结果
                msg = f"✅ 导入完成！\n\n成功: {imported} 条"
                if failed > 0:
                    msg += f"\n失败: {failed} 条"

                if failed_rows:
                    msg += f"\n\n前5个错误:\n" + "\n".join(failed_rows[:5])
                    if len(failed_rows) > 5:
                        msg += f"\n...还有 {len(failed_rows) - 5} 个错误"

                messagebox.showinfo("导入结果", msg)

            btn_frame = tk.Frame(preview_window)
            btn_frame.pack(pady=15)

            tk.Button(
                btn_frame,
                text="✅ 确认导入",
                command=do_import,
                font=('微软雅黑', 12, 'bold'),
                bg='#27ae60',
                fg='white',
                width=15
            ).pack(side=tk.LEFT, padx=10)

            tk.Button(
                btn_frame,
                text="❌ 取消",
                command=preview_window.destroy,
                font=('微软雅黑', 12),
                bg='#e74c3c',
                fg='white',
                width=15
            ).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            messagebox.showerror("错误", f"读取Excel失败: {str(e)}")

    def show_monthly_stats(self):
        """显示月度统计"""
        # 创建弹窗
        popup = tk.Toplevel(self.root)
        popup.title("月度统计")
        popup.geometry("550x500")
        popup.transient(self.root)

        # 选择月份区域
        select_frame = tk.Frame(popup)
        select_frame.pack(pady=15)

        tk.Label(select_frame, text="选择月份:", font=('微软雅黑', 12, 'bold')).pack(side=tk.LEFT, padx=5)

        # 年份下拉框
        year_var = tk.StringVar(value=str(datetime.now().year))
        year_values = [str(y) for y in range(2020, 2031)]
        year_combo = ttk.Combobox(select_frame, textvariable=year_var, values=year_values,
                                   width=6, font=('微软雅黑', 11), state='readonly')
        year_combo.pack(side=tk.LEFT, padx=5)
        tk.Label(select_frame, text="年", font=('微软雅黑', 11)).pack(side=tk.LEFT)

        # 月份下拉框
        month_var = tk.StringVar(value=str(datetime.now().month).zfill(2))
        month_values = [str(m).zfill(2) for m in range(1, 13)]
        month_combo = ttk.Combobox(select_frame, textvariable=month_var, values=month_values,
                                    width=4, font=('微软雅黑', 11), state='readonly')
        month_combo.pack(side=tk.LEFT, padx=5)
        tk.Label(select_frame, text="月", font=('微软雅黑', 11)).pack(side=tk.LEFT)

        # 确认按钮
        def on_confirm():
            calculate_stats()

        tk.Button(select_frame, text="确认查看", command=on_confirm,
                  font=('微软雅黑', 11, 'bold'), bg='#3498db', fg='white').pack(side=tk.LEFT, padx=15)

        # 分割线
        tk.Frame(popup, height=2, bg='#bdc3c7').pack(fill=tk.X, padx=20)

        # 结果显示区域
        result_text = tk.Text(popup, font=('微软雅黑', 11), height=18, width=55)
        result_text.pack(padx=20, pady=15, fill=tk.BOTH, expand=True)

        # 初始提示
        result_text.insert('1.0', '\n\n请选择年月后点击"确认查看"\n统计该月份的销售额和退货情况')

        def calculate_stats():
            year = year_var.get()
            month = month_var.get().zfill(2)
            year_month = f"{year}-{month}"
            month_records = [r for r in self.records if r['date'].startswith(year_month)]
            
            # 分离销售和退货
            sale_records = [r for r in month_records if r.get('type') != 'return' and r['quantity'] > 0]
            return_records = [r for r in month_records if r.get('type') == 'return' or r['quantity'] < 0]
            
            # 销售统计
            sale_qty = sum(r['quantity'] for r in sale_records)
            sale_amount = sum(r['total_amount'] for r in sale_records)
            
            # 退货统计
            return_qty = sum(abs(r['quantity']) for r in return_records)
            return_amount = sum(abs(r['total_amount']) for r in return_records)
            
            # 净统计
            net_qty = sale_qty - return_qty
            net_amount = sale_amount - return_amount
            
            avg_price = sale_amount / sale_qty if sale_qty > 0 else 0
            
            # 按日期分组
            daily_stats = {}
            for r in month_records:
                date = r['date']
                if date not in daily_stats:
                    daily_stats[date] = {'qty': 0, 'amount': 0, 'return_qty': 0, 'return_amount': 0}
                
                if r.get('type') == 'return' or r['quantity'] < 0:
                    daily_stats[date]['return_qty'] += abs(r['quantity'])
                    daily_stats[date]['return_amount'] += abs(r['total_amount'])
                else:
                    daily_stats[date]['qty'] += r['quantity']
                    daily_stats[date]['amount'] += r['total_amount']
            
            result = f"""
📊 {year_month} 月度统计
━━━━━━━━━━━━━━━━━━━━━━━━
✅ 销售: {sale_qty}套 ¥{sale_amount:.2f}
🔄 退货: {return_qty}套 ¥{return_amount:.2f}
━━━━━━━━━━━━━━━━━━━━━━━━
📦 净数量: {net_qty} 套
💵 净金额: ¥{net_amount:.2f}
💰 平均单价: ¥{avg_price:.2f}
📝 记录数: {len(month_records)} 条
📅 有记录天数: {len(daily_stats)} 天

📈 每日明细:
━━━━━━━━━━━━━━━━━━━━━━━━
"""
            for date in sorted(daily_stats.keys()):
                stats = daily_stats[date]
                day_result = f"{date}: "
                if stats['qty'] > 0:
                    day_result += f"售{stats['qty']}套¥{stats['amount']:.0f}"
                if stats['return_qty'] > 0:
                    if stats['qty'] > 0:
                        day_result += " | "
                    day_result += f"退{stats['return_qty']}套¥{stats['return_amount']:.0f}"
                result += day_result + "\n"
            
            result_text.delete('1.0', tk.END)
            result_text.insert('1.0', result)
        
        # 提示用户操作
        tk.Label(popup, text="（点击 确认查看 按钮刷新统计）", font=('微软雅黑', 9), fg='#7f8c8d').pack(pady=5)

    def show_settings(self):
        """显示系统设置窗口"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("⚙️ 系统设置")
        settings_window.geometry("450x400")
        settings_window.transient(self.root)
        settings_window.grab_set()
        
        # 标题
        tk.Label(settings_window, text="⚙️ 系统设置", font=('微软雅黑', 16, 'bold')).pack(pady=15)
        
        # 版本信息区
        version_frame = tk.LabelFrame(settings_window, text="版本信息", font=('微软雅黑', 11, 'bold'))
        version_frame.pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(version_frame, text=f"当前版本: v{VERSION}", font=('微软雅黑', 12)).pack(anchor='w', padx=15, pady=5)
        tk.Label(version_frame, text=f"GitHub仓库: {GITHUB_REPO}", font=('微软雅黑', 10), fg='#7f8c8d').pack(anchor='w', padx=15, pady=2)
        
        # 升级状态显示
        self.upgrade_status_var = tk.StringVar(value="点击下方按钮检查更新")
        status_label = tk.Label(version_frame, textvariable=self.upgrade_status_var, font=('微软雅黑', 10), fg='#2c3e50')
        status_label.pack(anchor='w', padx=15, pady=5)
        
        # 升级按钮
        btn_frame = tk.Frame(version_frame)
        btn_frame.pack(pady=10)
        
        check_btn = tk.Button(
            btn_frame,
            text="🔍 检查更新",
            command=lambda: self.check_for_updates(settings_window),
            font=('微软雅黑', 11),
            bg='#3498db',
            fg='white',
            width=12
        )
        check_btn.pack(side=tk.LEFT, padx=5)
        
        # 数据管理区
        data_frame = tk.LabelFrame(settings_window, text="数据管理", font=('微软雅黑', 11, 'bold'))
        data_frame.pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(data_frame, text=f"数据文件: {self.data_file}", font=('微软雅黑', 9), fg='#7f8c8d', wraplength=380).pack(anchor='w', padx=15, pady=5)
        tk.Label(data_frame, text=f"记录总数: {len(self.records)} 条", font=('微软雅黑', 10)).pack(anchor='w', padx=15, pady=2)
        
        # 打开数据目录按钮
        tk.Button(
            data_frame,
            text="📁 打开数据目录",
            command=lambda: os.startfile(self.data_dir) if os.name == 'nt' else None,
            font=('微软雅黑', 10),
            bg='#95a5a6',
            fg='white'
        ).pack(pady=10)
        
        # 关于区
        about_frame = tk.LabelFrame(settings_window, text="关于", font=('微软雅黑', 11, 'bold'))
        about_frame.pack(fill=tk.X, padx=20, pady=10)
        
        about_text = """家纺四件套记账系统
专为家纺销售设计的简单记账工具
支持多商品录入、退货管理、数据导入导出

© 2026 All Rights Reserved"""
        tk.Label(about_frame, text=about_text, font=('微软雅黑', 9), fg='#7f8c8d', justify=tk.LEFT).pack(padx=15, pady=10)
        
        # 关闭按钮
        tk.Button(
            settings_window,
            text="关闭",
            command=settings_window.destroy,
            font=('微软雅黑', 11),
            bg='#e74c3c',
            fg='white',
            width=10
        ).pack(pady=15)
    
    def check_for_updates(self, parent_window=None):
        """检查GitHub Release更新"""
        self.upgrade_status_var.set("⏳ 正在检查更新...")
        
        def check_thread():
            try:
                # 构建API URL
                api_url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
                
                req = urllib.request.Request(api_url)
                req.add_header('User-Agent', 'AccountingApp')
                
                with urllib.request.urlopen(req, timeout=10) as response:
                    data = json.loads(response.read().decode('utf-8'))
                
                latest_version = data.get('tag_name', '').lstrip('v')
                release_url = data.get('html_url', '')
                release_notes = data.get('body', '无更新说明')[:200]
                
                # 获取资源列表（用于自动更新）
                assets = data.get('assets', [])
                
                # 比较版本
                if self.compare_versions(latest_version, VERSION) > 0:
                    self.root.after(0, lambda: self.show_update_available(latest_version, release_url, release_notes, parent_window, assets))
                else:
                    self.root.after(0, lambda: self.upgrade_status_var.set(f"✅ 已是最新版本 v{VERSION}"))
                    
            except urllib.error.HTTPError as e:
                if e.code == 404:
                    self.root.after(0, lambda: self.upgrade_status_var.set("⚠️ 未找到发布版本，请先在GitHub创建Release"))
                else:
                    self.root.after(0, lambda: self.upgrade_status_var.set(f"❌ 检查失败: HTTP {e.code}"))
            except Exception as e:
                self.root.after(0, lambda: self.upgrade_status_var.set(f"❌ 检查失败: {str(e)[:30]}"))
        
        # 在后台线程中检查
        thread = threading.Thread(target=check_thread, daemon=True)
        thread.start()
    
    def compare_versions(self, v1, v2):
        """比较版本号，v1 > v2 返回 1，v1 < v2 返回 -1，相等返回 0"""
        def parse_version(v):
            return [int(x) for x in v.split('.') if x.isdigit()]
        
        parts1 = parse_version(v1)
        parts2 = parse_version(v2)
        
        # 补齐长度
        max_len = max(len(parts1), len(parts2))
        parts1.extend([0] * (max_len - len(parts1)))
        parts2.extend([0] * (max_len - len(parts2)))
        
        for p1, p2 in zip(parts1, parts2):
            if p1 > p2:
                return 1
            elif p1 < p2:
                return -1
        return 0
    
    def show_update_available(self, new_version, release_url, release_notes, parent_window=None, assets=None):
        """显示有新版本可用，并提供自动更新选项"""
        self.upgrade_status_var.set(f"🎉 发现新版本 v{new_version}!")

        # 检查是否在运行打包后的EXE
        is_exe = getattr(sys, 'frozen', False)

        # 创建更新对话框
        update_window = tk.Toplevel(parent_window or self.root)
        update_window.title("🎉 发现新版本")
        update_window.geometry("450x350")
        update_window.transient(parent_window or self.root)
        update_window.grab_set()

        tk.Label(update_window, text="🎉 发现新版本!", font=('微软雅黑', 14, 'bold'), fg='#27ae60').pack(pady=15)

        tk.Label(update_window, text=f"当前版本: v{VERSION}", font=('微软雅黑', 11)).pack()
        tk.Label(update_window, text=f"最新版本: v{new_version}", font=('微软雅黑', 11, 'bold'), fg='#3498db').pack()

        # 更新说明
        tk.Label(update_window, text="更新说明:", font=('微软雅黑', 10, 'bold')).pack(anchor='w', padx=20, pady=(15, 5))

        notes_text = tk.Text(update_window, font=('微软雅黑', 9), height=5, width=48, wrap=tk.WORD)
        notes_text.pack(padx=20)
        notes_text.insert('1.0', release_notes if release_notes else "暂无更新说明")
        notes_text.config(state='disabled')

        # 按钮
        btn_frame = tk.Frame(update_window)
        btn_frame.pack(pady=15)

        def open_release():
            import webbrowser
            webbrowser.open(release_url)

        # 如果是EXE运行，并且找到了EXE资源，显示自动更新按钮
        exe_asset = None
        if assets and is_exe:
            for asset in assets:
                name = asset.get('name', '').lower()
                # 检查是否是EXE文件（支持带版本号或不带版本号的文件名）
                if name.endswith('.exe'):
                    exe_asset = asset
                    break

        if exe_asset and is_exe:
            # 自动更新按钮
            def start_auto_update():
                update_window.destroy()
                self.auto_update_exe(exe_asset.get('browser_download_url'), new_version)

            tk.Button(
                btn_frame,
                text="⬇️ 自动更新",
                command=start_auto_update,
                font=('微软雅黑', 11),
                bg='#27ae60',
                fg='white',
                width=12
            ).pack(side=tk.LEFT, padx=5)
        else:
            # 手动下载按钮
            tk.Button(
                btn_frame,
                text="🌐 前往下载",
                command=open_release,
                font=('微软雅黑', 11),
                bg='#27ae60',
                fg='white',
                width=12
            ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btn_frame,
            text="稍后再说",
            command=update_window.destroy,
            font=('微软雅黑', 11),
            bg='#95a5a6',
            fg='white',
            width=10
        ).pack(side=tk.LEFT, padx=5)

    # ==================== 小票打印相关方法 ====================

    def print_receipt(self, record):
        """打印小票（包含关联的退货记录）"""
        try:
            # 获取紧凑模式设置（默认True）
            compact_mode = self.printer_settings.get('compact_mode', True)
            
            # 查找关联的退货记录
            return_records = self.get_return_records(record)
            
            # 生成小票文本（包含退货记录）
            receipt_text = self.receipt_printer.format_receipt(record, compact=compact_mode, return_records=return_records)

            # 获取用户选择的打印机
            printer_name = self.printer_settings.get('printer_name', '')

            # 打印到Windows打印机
            result = self.receipt_printer.print_to_windows_printer(receipt_text, printer_name if printer_name else None)

            if result['success']:
                messagebox.showinfo("打印成功", result['message'])
            else:
                # 打印失败，提供保存为文本的选项
                if messagebox.askyesno("打印失败", f"{result['message']}\n\n是否将小票保存为文本文件？"):
                    self.save_receipt_as_text(record)
        except Exception as e:
            messagebox.showerror("打印错误", f"打印失败: {str(e)}")

    def get_return_records(self, record):
        """获取关联的退货记录"""
        return_records = []
        is_return = record.get('type') == 'return' or record.get('quantity', 0) < 0
        if not is_return:
            # 是销售记录，查找关联的退货
            record_id = record.get('id')
            for r in self.records:
                if r.get('type') == 'return' or r.get('quantity', 0) < 0:
                    if r.get('original_record_id') == record_id:
                        return_records.append(r)
        return return_records
    
    def save_receipt_as_text(self, record):
        """保存小票为文本文件（包含退货记录）"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            initialfile=f"小票_{record.get('id', '0000')}_{record.get('date', '')}.txt"
        )

        if file_path:
            try:
                # 获取紧凑模式设置（默认True）
                compact_mode = self.printer_settings.get('compact_mode', True)
                # 查找关联的退货记录
                return_records = self.get_return_records(record)
                receipt_text = self.receipt_printer.format_receipt(record, compact=compact_mode, return_records=return_records)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(receipt_text)
                messagebox.showinfo("保存成功", f"小票已保存到:\n{file_path}")
            except Exception as e:
                messagebox.showerror("保存失败", f"保存失败: {str(e)}")

    def show_receipt_preview(self, record):
        """显示小票预览（包含退货记录）"""
        preview_window = tk.Toplevel(self.root)
        preview_window.title("小票预览")
        preview_window.geometry("450x650")
        preview_window.transient(self.root)

        # 标题
        tk.Label(
            preview_window,
            text="🧾 小票预览",
            font=('微软雅黑', 14, 'bold')
        ).pack(pady=10)

        # 创建文本框显示小票内容
        text_frame = tk.Frame(preview_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 获取紧凑模式设置（默认True）
        compact_mode = self.printer_settings.get('compact_mode', True)
        # 查找关联的退货记录
        return_records = self.get_return_records(record)
        receipt_text = self.receipt_printer.format_receipt(record, compact=compact_mode, return_records=return_records)

        text_widget = tk.Text(
            text_frame,
            font=('Courier New', 10),
            width=40,
            height=25,
            yscrollcommand=scrollbar.set,
            wrap=tk.NONE
        )
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)

        text_widget.insert('1.0', receipt_text)
        text_widget.config(state='disabled')

        # 按钮区域
        btn_frame = tk.Frame(preview_window)
        btn_frame.pack(pady=15)

        tk.Button(
            btn_frame,
            text="🖨️ 打印",
            command=lambda: [preview_window.destroy(), self.print_receipt(record)],
            font=('微软雅黑', 11),
            bg='#3498db',
            fg='white',
            width=10
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btn_frame,
            text="💾 保存",
            command=lambda: self.save_receipt_as_text(record),
            font=('微软雅黑', 11),
            bg='#9b59b6',
            fg='white',
            width=10
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btn_frame,
            text="关闭",
            command=preview_window.destroy,
            font=('微软雅黑', 11),
            bg='#95a5a6',
            fg='white',
            width=10
        ).pack(side=tk.LEFT, padx=5)

    def show_printer_settings(self):
        """显示打印设置窗口"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("🖨️ 打印设置")
        settings_window.geometry("500x650")
        settings_window.transient(self.root)
        settings_window.grab_set()

        # 创建Canvas和Scrollbar以支持滚动
        canvas = tk.Canvas(settings_window)
        scrollbar = tk.Scrollbar(settings_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 标题
        tk.Label(
            scrollable_frame,
            text="🖨️ 小票打印设置",
            font=('微软雅黑', 16, 'bold')
        ).pack(pady=15)

        # ========== 打印机选择区域 ==========
        printer_frame = tk.LabelFrame(
            scrollable_frame,
            text="打印机设置",
            font=('微软雅黑', 11, 'bold')
        )
        printer_frame.pack(fill=tk.X, padx=20, pady=10)

        # 获取打印机列表
        printers = get_printer_list()
        printer_choices = ["使用系统默认打印机"] + printers

        tk.Label(printer_frame, text="选择打印机:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky='w', padx=10, pady=5)
        printer_var = tk.StringVar(value=self.printer_settings.get('printer_name', '') or "使用系统默认打印机")
        printer_combo = ttk.Combobox(printer_frame, values=printer_choices, width=30, font=('微软雅黑', 10), state='readonly')
        printer_combo.set(printer_var.get())
        printer_combo.grid(row=0, column=1, padx=10, pady=5)

        # 刷新打印机列表按钮
        def refresh_printers():
            printers = get_printer_list()
            printer_choices = ["使用系统默认打印机"] + printers
            printer_combo['values'] = printer_choices
            messagebox.showinfo("提示", f"找到 {len(printers)} 台打印机")

        tk.Button(printer_frame, text="🔄 刷新", command=refresh_printers,
                  font=('微软雅黑', 9), bg='#3498db', fg='white').grid(row=0, column=2, padx=5, pady=5)

        # 纸张宽度选择
        tk.Label(printer_frame, text="纸张宽度:", font=('微软雅黑', 10)).grid(row=1, column=0, sticky='w', padx=10, pady=5)
        paper_width_var = tk.IntVar(value=self.printer_settings.get('paper_width', 58))
        paper_frame = tk.Frame(printer_frame)
        paper_frame.grid(row=1, column=1, sticky='w', padx=10, pady=5)
        tk.Radiobutton(paper_frame, text="58mm", variable=paper_width_var, value=58, font=('微软雅黑', 10)).pack(side=tk.LEFT, padx=5)
        tk.Radiobutton(paper_frame, text="76mm", variable=paper_width_var, value=76, font=('微软雅黑', 10)).pack(side=tk.LEFT, padx=5)
        tk.Radiobutton(paper_frame, text="80mm", variable=paper_width_var, value=80, font=('微软雅黑', 10)).pack(side=tk.LEFT, padx=5)

        # 自动打印选项
        auto_print_var = tk.BooleanVar(value=self.printer_settings.get('auto_print', False))
        tk.Checkbutton(printer_frame, text="销售/退货后自动打印小票", variable=auto_print_var,
                       font=('微软雅黑', 10)).grid(row=2, column=0, columnspan=2, sticky='w', padx=10, pady=5)

        # 紧凑模式选项
        compact_mode_var = tk.BooleanVar(value=self.printer_settings.get('compact_mode', True))
        tk.Checkbutton(printer_frame, text="紧凑模式（一张纸打印，推荐58mm/76mm）", variable=compact_mode_var,
                       font=('微软雅黑', 10)).grid(row=3, column=0, columnspan=2, sticky='w', padx=10, pady=5)

        # ========== 店铺信息区域 ==========
        shop_frame = tk.LabelFrame(
            scrollable_frame,
            text="店铺信息",
            font=('微软雅黑', 11, 'bold')
        )
        shop_frame.pack(fill=tk.X, padx=20, pady=10)

        # 店名
        tk.Label(shop_frame, text="店铺名称:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky='w', padx=10, pady=5)
        shop_name_var = tk.StringVar(value=self.printer_settings.get('shop_name', '家纺四件套'))
        tk.Entry(shop_frame, textvariable=shop_name_var, font=('微软雅黑', 10), width=35).grid(row=0, column=1, padx=10, pady=5)

        # 地址
        tk.Label(shop_frame, text="店铺地址:", font=('微软雅黑', 10)).grid(row=1, column=0, sticky='w', padx=10, pady=5)
        shop_address_var = tk.StringVar(value=self.printer_settings.get('shop_address', ''))
        tk.Entry(shop_frame, textvariable=shop_address_var, font=('微软雅黑', 10), width=35).grid(row=1, column=1, padx=10, pady=5)

        # 电话
        tk.Label(shop_frame, text="联系电话:", font=('微软雅黑', 10)).grid(row=2, column=0, sticky='w', padx=10, pady=5)
        shop_phone_var = tk.StringVar(value=self.printer_settings.get('shop_phone', ''))
        tk.Entry(shop_frame, textvariable=shop_phone_var, font=('微软雅黑', 10), width=35).grid(row=2, column=1, padx=10, pady=5)

        # ========== 小票样式区域 ==========
        style_frame = tk.LabelFrame(
            scrollable_frame,
            text="小票样式",
            font=('微软雅黑', 11, 'bold')
        )
        style_frame.pack(fill=tk.X, padx=20, pady=10)

        # 底部文字
        tk.Label(style_frame, text="底部文字:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky='w', padx=10, pady=5)
        footer_var = tk.StringVar(value=self.printer_settings.get('footer_text', '谢谢惠顾，欢迎下次光临！'))
        tk.Entry(style_frame, textvariable=footer_var, font=('微软雅黑', 10), width=35).grid(row=0, column=1, padx=10, pady=5)

        # ========== 预览区域 ==========
        preview_frame = tk.LabelFrame(
            scrollable_frame,
            text="小票预览",
            font=('微软雅黑', 11, 'bold')
        )
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        preview_text = tk.Text(preview_frame, font=('Courier New', 9), width=50, height=12)
        preview_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        def update_preview():
            """更新预览"""
            test_record = {
                "id": 8888,
                "date": datetime.now().strftime("%Y-%m-%d"),
                "created_at": datetime.now().strftime("%H:%M:%S"),
                "quantity": 2,
                "unit_price": 280.00,
                "total_amount": 560.00,
                "note": "测试预览",
                "items": [
                    {"quantity": 2, "unit_price": 280.00}
                ]
            }

            # 临时更新设置
            temp_printer = ReceiptPrinter()
            temp_printer.receipt_width = 32 if paper_width_var.get() == 58 else (44 if paper_width_var.get() == 76 else 48)
            temp_printer.set_shop_info(
                name=shop_name_var.get(),
                address=shop_address_var.get(),
                phone=shop_phone_var.get()
            )
            temp_printer.footer_text = footer_var.get()

            preview_text.delete('1.0', tk.END)
            # 使用紧凑模式预览
            preview_text.insert('1.0', temp_printer.format_receipt(test_record, compact=compact_mode_var.get()))

        # 初始预览
        update_preview()

        # 预览按钮
        tk.Button(
            scrollable_frame,
            text="👁️ 更新预览",
            command=update_preview,
            font=('微软雅黑', 10),
            bg='#9b59b6',
            fg='white'
        ).pack(pady=5)

        # ========== 按钮区域 ==========
        btn_frame = tk.Frame(scrollable_frame)
        btn_frame.pack(pady=15)

        def save_settings():
            """保存设置"""
            selected_printer = printer_combo.get()
            printer_name = "" if selected_printer == "使用系统默认打印机" else selected_printer

            settings = {
                'shop_name': shop_name_var.get(),
                'shop_address': shop_address_var.get(),
                'shop_phone': shop_phone_var.get(),
                'footer_text': footer_var.get(),
                'printer_name': printer_name,
                'paper_width': paper_width_var.get(),
                'auto_print': auto_print_var.get(),
                'compact_mode': compact_mode_var.get()
            }

            # 更新当前打印机设置
            self.receipt_printer.receipt_width = 32 if paper_width_var.get() == 58 else (44 if paper_width_var.get() == 76 else 48)
            self.receipt_printer.set_shop_info(
                name=settings['shop_name'],
                address=settings['shop_address'],
                phone=settings['shop_phone']
            )
            self.receipt_printer.footer_text = settings['footer_text']

            # 保存到文件
            self.save_printer_settings(settings)

            messagebox.showinfo("保存成功", "打印设置已保存！\n设置将在下次打印时生效。")
            settings_window.destroy()

        tk.Button(
            btn_frame,
            text="✅ 保存设置",
            command=save_settings,
            font=('微软雅黑', 11),
            bg='#27ae60',
            fg='white',
            width=12
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btn_frame,
            text="❌ 取消",
            command=settings_window.destroy,
            font=('微软雅黑', 11),
            bg='#e74c3c',
            fg='white',
            width=10
        ).pack(side=tk.LEFT, padx=5)

        # 布局Canvas和Scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def print_selected_record(self):
        """打印选中的记录"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要打印的记录")
            return

        item = selected[0]
        values = self.tree.item(item, 'values')
        record_id = int(values[0])

        # 找到对应记录
        record = None
        for r in self.records:
            if r['id'] == record_id:
                record = r
                break

        if record:
            self.show_receipt_preview(record)
        else:
            messagebox.showerror("错误", "未找到记录")

    def auto_update_exe(self, download_url, new_version):
        """自动下载并更新EXE文件"""
        # 创建下载进度窗口
        progress_window = tk.Toplevel(self.root)
        progress_window.title("⬇️ 正在下载更新")
        progress_window.geometry("400x200")
        progress_window.transient(self.root)
        progress_window.grab_set()

        tk.Label(progress_window, text=f"正在下载 v{new_version}...", font=('微软雅黑', 12)).pack(pady=20)

        progress_var = tk.DoubleVar(value=0)
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100, length=350)
        progress_bar.pack(pady=10, padx=20)

        status_label = tk.Label(progress_window, text="准备下载...", font=('微软雅黑', 10))
        status_label.pack(pady=10)

        def download_thread():
            try:
                # 获取当前EXE路径和目录
                current_exe = sys.executable
                exe_dir = os.path.dirname(current_exe)

                # 创建临时下载路径
                temp_dir = os.path.join(self.data_dir, 'update')
                os.makedirs(temp_dir, exist_ok=True)
                downloaded_exe = os.path.join(temp_dir, 'new_version.exe')

                # 下载文件
                status_label.config(text="正在下载...")

                req = urllib.request.Request(download_url)
                req.add_header('User-Agent', 'AccountingApp')

                with urllib.request.urlopen(req, timeout=120) as response:
                    total_size = int(response.headers.get('Content-Length', 0))
                    downloaded = 0
                    chunk_size = 8192

                    with open(downloaded_exe, 'wb') as f:
                        while True:
                            chunk = response.read(chunk_size)
                            if not chunk:
                                break
                            f.write(chunk)
                            downloaded += len(chunk)
                            if total_size > 0:
                                progress = (downloaded / total_size) * 100
                                self.root.after(0, lambda p=progress: progress_var.set(p))

                # 下载完成
                self.root.after(0, progress_window.destroy)
                self.root.after(0, lambda: self.install_update(downloaded_exe, current_exe, exe_dir, new_version))

            except Exception as e:
                self.root.after(0, progress_window.destroy)
                self.root.after(0, lambda: messagebox.showerror("下载失败", f"更新下载失败: {str(e)}\n\n请手动前往GitHub下载最新版本。"))
                self.root.after(0, lambda: self.upgrade_status_var.set("❌ 更新下载失败"))

        # 启动下载线程
        thread = threading.Thread(target=download_thread, daemon=True)
        thread.start()

    def install_update(self, downloaded_exe, current_exe, exe_dir, new_version):
        """安装更新：创建批处理脚本替换旧版本"""
        # 确认安装
        if not messagebox.askyesno("安装更新", f"新版本 v{new_version} 已下载完成！\n\n点击「是」关闭程序并安装更新\n程序将自动重启"):
            self.upgrade_status_var.set("⏸️ 更新已取消")
            return

        try:
            # 使用固定的新文件名（不带版本号，方便后续更新）
            new_exe_name = "家纺记账系统.exe"
            new_exe_path = os.path.join(exe_dir, new_exe_name)
            old_exe_backup = os.path.join(exe_dir, "家纺记账系统.exe.old")
            current_backup = current_exe + ".old"

            # 创建批处理脚本
            batch_path = os.path.join(self.data_dir, 'update.bat')

            batch_content = f'''@echo off
chcp 65001 >nul
echo ==========================================
echo   正在安装更新 v{new_version}...
echo ==========================================
echo.
timeout /t 2 /nobreak >nul

REM 等待原程序完全退出
echo 等待程序退出...
timeout /t 3 /nobreak >nul

REM 删除旧备份文件（如果存在）
if exist "{current_backup}" del /f "{current_backup}" 2>nul
if exist "{old_exe_backup}" del /f "{old_exe_backup}" 2>nul

REM 备份当前EXE（重命名为 .old）
echo 备份当前版本...
ren "{current_exe}" "{os.path.basename(current_backup)}" 2>nul

REM 将新版本复制为固定文件名
echo 安装新版本...
copy /y "{downloaded_exe}" "{new_exe_path}" >nul

REM 检查复制是否成功
if not exist "{new_exe_path}" (
    echo 安装失败，恢复旧版本...
    ren "{current_backup}" "{os.path.basename(current_exe)}" 2>nul
    echo 更新失败，请手动下载安装。
    pause
    exit /b 1
)

REM 启动新版本
echo.
echo ==========================================
echo   启动新版本...
echo ==========================================
timeout /t 1 /nobreak >nul
start "" "{new_exe_path}"

REM 清理临时文件
del /f "{downloaded_exe}" 2>nul

REM 删除自己
del /f "%~f0"
'''

            with open(batch_path, 'w', encoding='utf-8') as f:
                f.write(batch_content)

            # 执行批处理并退出
            subprocess.Popen(batch_path, shell=True)
            self.root.quit()

        except Exception as e:
            messagebox.showerror("安装失败", f"更新安装失败: {str(e)}\n\n请手动前往GitHub下载最新版本。")
            self.upgrade_status_var.set("❌ 更新安装失败")


def main():
    """主程序"""
    root = tk.Tk()
    
    # 设置样式
    style = ttk.Style()
    style.theme_use('clam')
    
    # 设置中文字体
    root.option_add('*Font', '微软雅黑 10')
    
    app = AccountingApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
